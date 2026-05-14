#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Report Generator 单元测试

测试内容:
1. 车辆信息表生成
2. 组件汇总表生成
3. 详细结果表生成
4. SOC提取和映射
5. 单位获取
6. 多格式车辆信息提取
"""

import sys
from pathlib import Path

import pytest
import pandas as pd
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

from generate_excel_report import (
    generate_excel_report,
    extract_vehicle_info_value,
    extract_soc_from_condition_id,
    get_soc_level,
    get_unit_from_component,
    create_vehicle_info_sheet,
    create_component_summary_sheet,
    create_detailed_results_sheet,
    load_json_data
)


class TestExtractVehicleInfoValue:
    """测试车辆信息值提取"""

    def test_extract_with_primary_key(self):
        """测试使用主键提取"""
        vehicle_info = {'车型': '坦克500', '制造商': '长城'}

        result = extract_vehicle_info_value(vehicle_info, ['车型'], [])
        assert result == '坦克500'

    def test_extract_with_fallback_key(self):
        """测试使用备用键提取"""
        vehicle_info = {'参数名称': 'iCAR V27'}

        result = extract_vehicle_info_value(vehicle_info, ['车型'], ['参数名称'])
        assert result == 'iCAR V27'

    def test_extract_priority_order(self):
        """测试优先级顺序"""
        vehicle_info = {'车型': '坦克500', '参数名称': 'iCAR V27'}

        # 主键优先
        result = extract_vehicle_info_value(vehicle_info, ['车型'], ['参数名称'])
        assert result == '坦克500'

    def test_extract_not_found(self):
        """测试未找到时返回空字符串"""
        vehicle_info = {'其他键': '值'}

        result = extract_vehicle_info_value(vehicle_info, ['车型'], [])
        assert result == ''

    def test_extract_empty_value(self):
        """测试空值处理"""
        vehicle_info = {'车型': '', '参数名称': '坦克500'}

        result = extract_vehicle_info_value(vehicle_info, ['车型'], ['参数名称'])
        # 空字符串应该被跳过，使用备用键
        assert result == '坦克500'


class TestExtractSocFromConditionId:
    """测试从条件ID提取SOC"""

    def test_standard_format(self):
        """测试标准格式"""
        assert extract_soc_from_condition_id('87_超车80-140(运动模式)') == 87
        assert extract_soc_from_condition_id('20_直流充电暖风') == 20
        assert extract_soc_from_condition_id('40_匀速80') == 40

    def test_slope_format(self):
        """测试坡度格式"""
        assert extract_soc_from_condition_id('坡度10_81_匀速80暖风') == 81
        assert extract_soc_from_condition_id('坡度10_32_急加速') == 32

    def test_invalid_format(self):
        """测试无效格式"""
        assert extract_soc_from_condition_id('') is None
        assert extract_soc_from_condition_id(None) is None
        assert extract_soc_from_condition_id('invalid') is None

    def test_non_numeric_prefix(self):
        """测试非数字前缀"""
        assert extract_soc_from_condition_id('abc_测试') is None

    def test_dash_separator(self):
        """测试-分隔符格式（V0006等车辆）"""
        assert extract_soc_from_condition_id('25-交流充电冷风') == 25
        assert extract_soc_from_condition_id('55-直流充电暖风') == 55
        assert extract_soc_from_condition_id('87-匀速100暖风（运动模式）') == 87
        assert extract_soc_from_condition_id('39-超车80-140（运动模式）dmd') == 39

    def test_slope_with_dash_separator(self):
        """测试坡度工况使用-分隔符"""
        assert extract_soc_from_condition_id('坡度10-24-匀速80暖风') == 24
        assert extract_soc_from_condition_id('坡度10-31-匀速80冷风（运动模式）') == 31

    def test_gbk_corruption(self):
        """测试GBK乱码坡度前缀"""
        assert extract_soc_from_condition_id('�¶�10_26_匀速80冷风') == 26
        assert extract_soc_from_condition_id('�¶�10_27_匀速80暖风') == 27
        assert extract_soc_from_condition_id('�¶�10-24-匀速80暖风（运动模式）') == 24

    def test_slope_with_space_separator(self):
        """测试坡度工况使用空格分隔符（V0009/V0010）"""
        assert extract_soc_from_condition_id('�¶�10 47_匀速80冷风') == 47
        assert extract_soc_from_condition_id('�¶�10 51_匀速80暖风') == 51
        assert extract_soc_from_condition_id('�¶�10 15_匀速80暖风') == 15

    def test_mixed_separators(self):
        """测试混用分隔符和边界情况"""
        assert extract_soc_from_condition_id('32_多次 加速') == 32
        assert extract_soc_from_condition_id('坡度100_测试') is None  # 负向前瞻保护
        assert extract_soc_from_condition_id('') is None
        assert extract_soc_from_condition_id(None) is None
        assert extract_soc_from_condition_id('invalid') is None


class TestGetSocLevel:
    """测试SOC等级映射"""

    def test_high_soc(self):
        """测试高电量"""
        assert get_soc_level(87) == '≥70%'
        assert get_soc_level(70) == '≥70%'
        assert get_soc_level(100) == '≥70%'

    def test_medium_soc(self):
        """测试中电量"""
        assert get_soc_level(50) == '40%-70%'
        assert get_soc_level(40) == '40%-70%'
        assert get_soc_level(69) == '40%-70%'

    def test_low_soc(self):
        """测试低电量"""
        assert get_soc_level(20) == '≤40%'
        assert get_soc_level(39) == '≤40%'
        assert get_soc_level(0) == '≤40%'

    def test_none_soc(self):
        """测试None值"""
        assert get_soc_level(None) == 'Unknown'


class TestGetUnitFromComponent:
    """测试从组件代码获取单位"""

    def test_voltage_component(self):
        """测试电压组件"""
        assert get_unit_from_component('FM_V') == 'V'
        assert get_unit_from_component('DC_V') == 'V'

    def test_current_component(self):
        """测试电流组件"""
        assert get_unit_from_component('FM_A') == 'A'
        assert get_unit_from_component('DCC_A') == 'A'

    def test_unknown_component(self):
        """测试未知组件"""
        assert get_unit_from_component('UNKNOWN') == ''
        assert get_unit_from_component('FM') == ''


class TestGenerateExcelReport:
    """测试生成完整Excel报告"""

    def test_generate_excel_report(self, tmp_path):
        """测试生成完整报告"""
        test_data = {
            'vehicle': {
                'vehicle_id': 'TEST01',
                'vehicle_info': {
                    '车型': '测试车',
                    '车长mm': '5000',
                    '制造商': '测试厂商'
                }
            },
            'components': {
                'FM_V': {
                    'component_name': '前电机电压',
                    'unit': 'V',
                    'conditions': {
                        '87_test': {
                            'condition_name': '测试工况',
                            'soc_level': '≥70%',
                            'time_domain': {'effective_value': 1.0, 'vpp': 0.5},
                            'frequency_domain': {
                                'peak_ranking': '1st',
                                'peak_frequency_khz': 10.0,
                                'peak_amplitude': 0.1,
                                'rms': 0.05
                            },
                            'image_path': '/path/to/image.png'
                        }
                    }
                }
            },
            'metadata': {
                'total_components': 1,
                'total_conditions': 1,
                'warnings': []
            }
        }

        output_path = tmp_path / "test_report.xlsx"
        generate_excel_report(test_data, str(output_path))

        assert output_path.exists()

        # 验证工作表
        xls = pd.ExcelFile(str(output_path))
        assert 'Vehicle Information' in xls.sheet_names
        assert 'Component Summary' in xls.sheet_names
        assert 'Detailed Results' in xls.sheet_names

    def test_generate_excel_empty_components(self, tmp_path):
        """测试空组件列表"""
        test_data = {
            'vehicle': {
                'vehicle_id': 'TEST01',
                'vehicle_info': {'车型': '测试车'}
            },
            'components': {},
            'metadata': {'total_components': 0, 'total_conditions': 0, 'warnings': []}
        }

        output_path = tmp_path / "test_report.xlsx"
        generate_excel_report(test_data, str(output_path))

        assert output_path.exists()

    def test_generate_excel_with_slope_conditions(self, tmp_path):
        """测试包含坡度工况"""
        test_data = {
            'vehicle': {
                'vehicle_id': 'TEST01',
                'vehicle_info': {'车型': '测试车'}
            },
            'components': {
                'FM_V': {
                    'component_name': '前电机电压',
                    'unit': 'V',
                    'conditions': {
                        '坡度10_81_匀速80暖风': {
                            'condition_name': '爬坡高温',
                            'soc_level': '≥70%',
                            'time_domain': {'effective_value': 2.0, 'vpp': 1.0},
                            'frequency_domain': {
                                'peak_ranking': '2nd',
                                'peak_frequency_khz': 15.0,
                                'peak_amplitude': 0.2,
                                'rms': 0.1
                            },
                            'image_path': '/path/to/image.png'
                        }
                    }
                }
            },
            'metadata': {'total_components': 1, 'total_conditions': 1, 'warnings': []}
        }

        output_path = tmp_path / "test_report.xlsx"
        generate_excel_report(test_data, str(output_path))

        assert output_path.exists()

        # 验证SOC等级正确提取
        df = pd.read_excel(str(output_path), sheet_name='Detailed Results')
        assert len(df) == 1
        assert df.iloc[0]['SOC Level'] == '≥70%'


class TestVehicleInfoSheet:
    """测试车辆信息表"""

    def test_vehicle_info_sheet_content(self, tmp_path):
        """测试车辆信息表内容"""
        from openpyxl import Workbook

        wb = Workbook()
        test_data = {
            'vehicle': {
                'vehicle_id': 'TEST01',
                'vehicle_info': {
                    '车型': '坦克500',
                    '制造商': '长城',
                    '车长mm': '5078',
                    '自定义字段': '自定义值'
                }
            }
        }

        create_vehicle_info_sheet(wb, test_data)

        ws = wb['Vehicle Information']
        assert ws['A1'].value == 'Vehicle Information'

    def test_vehicle_info_priority_fields(self, tmp_path):
        """测试优先字段"""
        from openpyxl import Workbook

        wb = Workbook()
        test_data = {
            'vehicle': {
                'vehicle_id': 'TEST01',
                'vehicle_info': {
                    '参数名称': 'iCAR V27',  # 备用字段
                    '长度(mm)': '5055',
                }
            }
        }

        create_vehicle_info_sheet(wb, test_data)

        ws = wb['Vehicle Information']
        # 验证Vehicle ID在第一行
        assert ws['A3'].value == 'Vehicle ID'
        assert ws['B3'].value == 'TEST01'


class TestComponentSummarySheet:
    """测试组件汇总表"""

    def test_component_summary_headers(self, tmp_path):
        """测试组件汇总表头"""
        from openpyxl import Workbook

        wb = Workbook()
        test_data = {
            'components': {
                'FM_V': {
                    'component_name': '前电机电压',
                    'unit': 'V',
                    'conditions': {'cond1': {}, 'cond2': {}}
                },
                'FM_A': {
                    'component_name': '前电机电流',
                    'unit': 'A',
                    'conditions': {'cond3': {}}
                }
            }
        }

        # 移除默认sheet
        wb.remove(wb.active)
        create_component_summary_sheet(wb, test_data)

        ws = wb['Component Summary']
        assert ws['A1'].value == 'Component Code'
        assert ws['B1'].value == 'Component Name'
        assert ws['C1'].value == 'Unit'
        assert ws['D1'].value == 'Conditions Count'

    def test_component_summary_data(self, tmp_path):
        """测试组件汇总数据"""
        from openpyxl import Workbook

        wb = Workbook()
        test_data = {
            'components': {
                'FM_V': {
                    'component_name': '前电机电压',
                    'unit': 'V',
                    'conditions': {'cond1': {}, 'cond2': {}}
                }
            }
        }

        # 移除默认sheet
        wb.remove(wb.active)
        create_component_summary_sheet(wb, test_data)

        ws = wb['Component Summary']
        assert ws['A2'].value == 'FM_V'
        assert ws['B2'].value == '前电机电压'
        assert ws['C2'].value == 'V'
        assert ws['D2'].value == 2


class TestDetailedResultsSheet:
    """测试详细结果表"""

    def test_detailed_results_headers(self, tmp_path):
        """测试详细结果表头"""
        from openpyxl import Workbook

        wb = Workbook()
        test_data = {
            'components': {}
        }

        # 移除默认sheet
        wb.remove(wb.active)
        create_detailed_results_sheet(wb, test_data)

        ws = wb['Detailed Results']
        expected_headers = ['No.', 'Component', 'Unit', 'Condition ID', 'Condition Name',
                           'SOC Level', 'Time Effective Value', 'Time VPP', 'Peak Ranking',
                           'Freq Peak (kHz)', 'Freq Peak Amplitude', 'Freq RMS', 'Image Path']

        for idx, header in enumerate(expected_headers, start=1):
            assert ws.cell(row=1, column=idx).value == header

    def test_detailed_results_sequence_numbering(self, tmp_path):
        """测试详细结果序号"""
        from openpyxl import Workbook

        wb = Workbook()
        test_data = {
            'components': {
                'FM_V': {
                    'component_name': '前电机电压',
                    'unit': 'V',
                    'conditions': {
                        '87_test1': {
                            'condition_name': '测试1',
                            'soc_level': '≥70%',
                            'time_domain': {'effective_value': 1.0, 'vpp': 0.5},
                            'frequency_domain': {
                                'peak_ranking': '1st',
                                'peak_frequency_khz': 10.0,
                                'peak_amplitude': 0.1,
                                'rms': 0.05
                            },
                            'image_path': '/path/1.png'
                        },
                        '20_test2': {
                            'condition_name': '测试2',
                            'soc_level': '≤40%',
                            'time_domain': {'effective_value': 2.0, 'vpp': 1.0},
                            'frequency_domain': {
                                'peak_ranking': '2nd',
                                'peak_frequency_khz': 20.0,
                                'peak_amplitude': 0.2,
                                'rms': 0.1
                            },
                            'image_path': '/path/2.png'
                        }
                    }
                }
            }
        }

        # 移除默认sheet
        wb.remove(wb.active)
        create_detailed_results_sheet(wb, test_data)

        ws = wb['Detailed Results']
        # 验证序号从1开始递增
        assert ws['A2'].value == 1
        assert ws['A3'].value == 2


class TestLoadJsonData:
    """测试加载JSON数据"""

    def test_load_valid_json(self, tmp_path):
        """测试加载有效JSON"""
        json_path = tmp_path / "test.json"
        test_data = {'vehicle': {'vehicle_id': 'TEST01'}, 'components': {}}

        with open(json_path, 'w', encoding='utf-8') as f:
            import json
            json.dump(test_data, f)

        result = load_json_data(str(json_path))
        assert result['vehicle']['vehicle_id'] == 'TEST01'

    def test_load_nonexistent_json(self, tmp_path):
        """测试加载不存在的JSON"""
        json_path = tmp_path / "nonexistent.json"

        with pytest.raises(FileNotFoundError):
            load_json_data(str(json_path))


class TestEdgeCases:
    """测试边界情况"""

    def test_missing_vehicle_info(self, tmp_path):
        """测试缺少车辆信息"""
        test_data = {
            'vehicle': {
                'vehicle_id': 'TEST01',
                'vehicle_info': {}
            },
            'components': {},
            'metadata': {'total_components': 0, 'total_conditions': 0, 'warnings': []}
        }

        output_path = tmp_path / "test_report.xlsx"
        # 不应该抛出异常
        generate_excel_report(test_data, str(output_path))
        assert output_path.exists()

    def test_missing_time_domain_data(self, tmp_path):
        """测试缺少时域数据"""
        test_data = {
            'vehicle': {
                'vehicle_id': 'TEST01',
                'vehicle_info': {'车型': '测试车'}
            },
            'components': {
                'FM_V': {
                    'component_name': '前电机电压',
                    'unit': 'V',
                    'conditions': {
                        '87_test': {
                            'condition_name': '测试工况',
                            'soc_level': '≥70%',
                            'time_domain': {},  # 空时域数据
                            'frequency_domain': {},  # 空频域数据
                            'image_path': ''
                        }
                    }
                }
            },
            'metadata': {'total_components': 1, 'total_conditions': 1, 'warnings': []}
        }

        output_path = tmp_path / "test_report.xlsx"
        generate_excel_report(test_data, str(output_path))
        assert output_path.exists()

    def test_very_long_strings(self, tmp_path):
        """测试超长字符串"""
        test_data = {
            'vehicle': {
                'vehicle_id': 'TEST01',
                'vehicle_info': {'车型': '测试车' * 100}
            },
            'components': {
                'FM_V': {
                    'component_name': '前电机电压',
                    'unit': 'V',
                    'conditions': {
                        '87_test': {
                            'condition_name': '测试工况' * 100,
                            'soc_level': '≥70%',
                            'time_domain': {'effective_value': 1.0, 'vpp': 0.5},
                            'frequency_domain': {
                                'peak_ranking': '1st',
                                'peak_frequency_khz': 10.0,
                                'peak_amplitude': 0.1,
                                'rms': 0.05
                            },
                            'image_path': '/path/' + 'a' * 500 + '.png'
                        }
                    }
                }
            },
            'metadata': {'total_components': 1, 'total_conditions': 1, 'warnings': []}
        }

        output_path = tmp_path / "test_report.xlsx"
        generate_excel_report(test_data, str(output_path))
        assert output_path.exists()


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
