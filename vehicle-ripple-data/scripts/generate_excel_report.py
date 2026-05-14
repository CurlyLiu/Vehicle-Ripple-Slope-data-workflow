#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vehicle High-Voltage Ripple Test Data - Excel Report Generator V4.0
Generate Excel report from vehicle ripple test data

Version: 4.0
Changes:
- Updated Detailed Results to 13 columns matching V0002 reference format
- Added Time Effective Value, Peak Ranking, Freq Peak Amplitude, Freq RMS columns
- Fixed Image Path to use absolute paths
- Standardized Excel format with English headers

Usage:
    python generate_excel_report.py --input-json data.json --output-excel report.xlsx
    python generate_excel_report.py --vehicle-folder V0001 --output-excel report.xlsx
"""

import json
import argparse
import os
import re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Import config manager
try:
    from config import ConfigManager, load_config
except ImportError:
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from config import ConfigManager, load_config

def load_json_data(json_path):
    """Load JSON data"""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

# 坡度前缀匹配正则（支持正常文本、GBK乱码、多种分隔符）
_SLOPE_PREFIX_PATTERN = re.compile(
    r'^(坡度|�¶�)\s*10(?![0-9])[_\-\s]*(\d+)[_\-\s]',
    re.IGNORECASE
)

# 普通工况SOC匹配正则（开头的数字 + 任意分隔符）
_SOC_PATTERN = re.compile(r'^(\d+)[_\-\s]')


def extract_soc_from_condition_id(condition_id):
    """
    Extract SOC value from condition_id

    Supports formats:
    - Standard: "20_直流充电暖风" -> 20
    - Dash: "25-交流充电冷风" -> 25
    - Slope: "坡度10_32_匀速80冷风" -> 32
    - Slope with dash: "坡度10-24-匀速80暖风" -> 24
    - Slope with space: "坡度10 47_匀速80冷风" -> 47
    - GBK corrupted: "�¶�10_26_匀速80冷风" -> 26
    """
    if not condition_id:
        return None

    # 坡度工况（支持GBK乱码和多种分隔符）
    slope_match = _SLOPE_PREFIX_PATTERN.match(condition_id)
    if slope_match:
        return int(slope_match.group(2))

    # 普通工况（开头的数字 + 任意分隔符）
    normal_match = _SOC_PATTERN.match(condition_id)
    if normal_match:
        return int(normal_match.group(1))

    return None

def get_soc_level(soc_value):
    """
    Map SOC value to SOC Level category
    
    Returns:
    - "≥70%" if SOC >= 70
    - "40%-70%" if 40 <= SOC < 70
    - "≤40%" if SOC < 40
    - "Unknown" if SOC is None
    """
    if soc_value is None:
        return "Unknown"
    
    if soc_value >= 70:
        return "≥70%"
    elif soc_value >= 40:
        return "40%-70%"
    else:
        return "≤40%"

def get_unit_from_component(component_code):
    """
    Get unit (A or V) from component code
    
    Component codes ending with _A are current (Amperes)
    Component codes ending with _V are voltage (Volts)
    """
    if component_code.endswith('_A'):
        return 'A'
    elif component_code.endswith('_V'):
        return 'V'
    else:
        return ''

def extract_vehicle_info_value(vehicle_info, primary_keys, fallback_keys=None):
    """
    Extract vehicle info value with multiple key mappings
    
    Supports different formats of vehicle_info files
    
    Args:
        vehicle_info: Dict containing vehicle information
        primary_keys: List of primary keys to try
        fallback_keys: Optional list of fallback keys
        
    Returns:
        The value if found, empty string otherwise
    """
    # Try primary keys first
    for key in primary_keys:
        if key in vehicle_info and vehicle_info[key]:
            return vehicle_info[key]
    
    # Try fallback keys
    if fallback_keys:
        for key in fallback_keys:
            if key in vehicle_info and vehicle_info[key]:
                return vehicle_info[key]
    
    return ''


def create_vehicle_info_sheet(wb, vehicle_data):
    """Create Vehicle Information sheet - Format matches slope-data skill (Parameter/Value columns)"""
    ws = wb.active
    ws.title = "Vehicle Information"

    # Vehicle info data
    vehicle = vehicle_data.get('vehicle', {})
    vehicle_info = vehicle.get('vehicle_info', {})

    # Add header row - "Parameter" and "Value" (matching slope-data format)
    header_fill = PatternFill(start_color='D5E8F0', end_color='D5E8F0', fill_type='solid')

    cell = ws.cell(row=1, column=1, value='Parameter')
    cell.font = Font(bold=True, color='000000')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = ws.cell(row=1, column=2, value='Value')
    cell.font = Font(bold=True, color='000000')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Build fields list (unified format with slope-data skill)
    # Standard fields (same order as slope-data)
    standard_fields = [
        ('Vehicle ID', vehicle.get('vehicle_id', 'Unknown')),
        ('Vehicle Model', vehicle_info.get('vehicle_model', 'Unknown')),
        ('Manufacturer', vehicle_info.get('manufacturer', 'Unknown')),
        ('Length (mm)', vehicle_info.get('length_mm', '')),
        ('Width (mm)', vehicle_info.get('width_mm', '')),
        ('Height (mm)', vehicle_info.get('height_mm', '')),
    ]

    # Add extra fields from vehicle_info (same exclusions as slope-data)
    excluded_keys = {
        'vehicle_id', 'vehicle_model', 'manufacturer',
        'length_mm', 'width_mm', 'height_mm',
            '---',  # separator row
    }

    extra_fields = []
    for key, value in vehicle_info.items():
        if key not in excluded_keys and value and str(value).strip():
            # Convert display name: title case for underscored keys
            if '_' in key and not any(c in key for c in '()（）'):
                display_key = key.replace('_', ' ').title()
            else:
                display_key = key
            extra_fields.append((display_key, value))

    # Combine all fields
    all_fields = standard_fields + extra_fields

    # Fill data starting from row 2 (after header)
    for row_idx, (key, value) in enumerate(all_fields, start=2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center')

    # Column widths - matching slope-data
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30

    return ws

def create_component_summary_sheet(wb, data):
    """Create Component Summary sheet"""
    ws = wb.create_sheet(title="Component Summary")
    
    # Headers
    headers = ['Component Code', 'Component Name', 'Unit', 'Conditions Count']
    header_fill = PatternFill(start_color='D5E8F0', end_color='D5E8F0', fill_type='solid')
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Write component data
    components = data.get('components', {})
    row_idx = 2
    
    for component_code, component_data in components.items():
        conditions = component_data.get('conditions', {})
        unit = get_unit_from_component(component_code)
        
        ws.cell(row=row_idx, column=1, value=component_code)
        ws.cell(row=row_idx, column=2, value=component_data.get('component_name', ''))
        ws.cell(row=row_idx, column=3, value=unit)
        ws.cell(row=row_idx, column=4, value=len(conditions))
        
        # Apply borders and alignment
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Arial', size=10)
        
        row_idx += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    
    return ws

def create_detailed_results_sheet(wb, data):
    """Create Detailed Results sheet (V4.0 format with 13 columns matching V0002 reference)"""
    ws = wb.create_sheet(title="Detailed Results")
    
    # Headers (V4.0 format with 13 columns matching V0002 reference)
    headers = ['No.', 'Component', 'Unit', 'Condition ID', 'Condition Name', 
               'SOC Level', 'Time Effective Value', 'Time VPP', 'Peak Ranking',
               'Freq Peak (kHz)', 'Freq Peak Amplitude', 'Freq RMS', 'Image Path']
    header_fill = PatternFill(start_color='D5E8F0', end_color='D5E8F0', fill_type='solid')
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Write detailed data
    components = data.get('components', {})
    row_idx = 2
    sequence_no = 1
    
    for component_code, component_data in components.items():
        component_name = component_data.get('component_name', '')
        unit = get_unit_from_component(component_code)
        conditions = component_data.get('conditions', {})
        
        for condition_id, condition_data in conditions.items():
            time_domain = condition_data.get('time_domain', {})
            freq_domain = condition_data.get('frequency_domain', {})
            
            # Extract SOC from condition_id (V3.0 fix)
            soc_value = extract_soc_from_condition_id(condition_id)
            soc_level = get_soc_level(soc_value)
            
            ws.cell(row=row_idx, column=1, value=sequence_no)  # No.
            ws.cell(row=row_idx, column=2, value=component_code)  # Component
            ws.cell(row=row_idx, column=3, value=unit)  # Unit
            ws.cell(row=row_idx, column=4, value=condition_id)  # Condition ID
            ws.cell(row=row_idx, column=5, value=condition_data.get('condition_name', ''))  # Condition Name
            ws.cell(row=row_idx, column=6, value=soc_level)  # SOC Level
            ws.cell(row=row_idx, column=7, value=time_domain.get('effective_value', ''))  # Time Effective Value
            ws.cell(row=row_idx, column=8, value=time_domain.get('vpp', ''))  # Time VPP
            ws.cell(row=row_idx, column=9, value=freq_domain.get('peak_ranking', ''))  # Peak Ranking
            ws.cell(row=row_idx, column=10, value=freq_domain.get('peak_frequency_khz', ''))  # Freq Peak
            ws.cell(row=row_idx, column=11, value=freq_domain.get('peak_amplitude', ''))  # Freq Peak Amplitude
            ws.cell(row=row_idx, column=12, value=freq_domain.get('rms', ''))  # Freq RMS
            ws.cell(row=row_idx, column=13, value=condition_data.get('image_path', ''))  # Image Path
            
            # Apply borders and alignment
            for col_idx in range(1, 14):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx in [5, 9, 13]:  # Text columns left align
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Arial', size=9)
            
            row_idx += 1
            sequence_no += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 8   # No.
    ws.column_dimensions['B'].width = 18  # Component
    ws.column_dimensions['C'].width = 8   # Unit
    ws.column_dimensions['D'].width = 28  # Condition ID
    ws.column_dimensions['E'].width = 30  # Condition Name
    ws.column_dimensions['F'].width = 12  # SOC Level
    ws.column_dimensions['G'].width = 20  # Time Effective Value
    ws.column_dimensions['H'].width = 12  # Time VPP
    ws.column_dimensions['I'].width = 40  # Peak Ranking
    ws.column_dimensions['J'].width = 15  # Freq Peak
    ws.column_dimensions['K'].width = 18  # Freq Peak Amplitude
    ws.column_dimensions['L'].width = 12  # Freq RMS
    ws.column_dimensions['M'].width = 60  # Image Path
    
    return ws

def generate_excel_report(data, output_path):
    """Generate complete Excel report (V3.0 format)"""
    wb = Workbook()
    
    # Create three sheets
    create_vehicle_info_sheet(wb, data)
    create_component_summary_sheet(wb, data)
    create_detailed_results_sheet(wb, data)
    
    # Save file
    wb.save(output_path)
    print(f"Excel report generated (V3.0 format): {output_path}")
    return output_path


class ConfigurableExcelGenerator:
    """配置驱动的Excel报表生成器"""
    
    def __init__(self, template_name='ripple/excel_template', skill_root=None):
        """
        初始化配置化报表生成器
        
        Args:
            template_name: 模板配置名称
            skill_root: Skill根目录
        """
        if skill_root is None:
            skill_root = Path(__file__).parent.parent
        
        self.config_mgr = ConfigManager(skill_root, hot_reload=True)
        self.template = self.config_mgr.load(template_name)
        self.styles = self.config_mgr.load('common/styles')
    
    def generate(self, data, output_path):
        """
        根据配置生成Excel报表
        
        Args:
            data: 处理后的数据字典
            output_path: 输出文件路径
            
        Returns:
            输出文件路径
        """
        wb = Workbook()
        
        # 删除默认sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 按顺序生成工作表
        sheets_config = sorted(self.template['sheets'], key=lambda x: x.get('order', 999))
        
        for sheet_config in sheets_config:
            self._create_sheet(wb, sheet_config, data)
        
        # 保存文件
        wb.save(output_path)
        print(f"配置化Excel报表已生成: {output_path}")
        return output_path
    
    def _create_sheet(self, wb, config, data):
        """根据配置创建单个工作表"""
        ws = wb.create_sheet(config['name'])
        
        layout = config.get('layout', 'table')
        
        if layout == 'vertical':
            self._create_vertical_sheet(ws, config, data)
        elif layout == 'table':
            self._create_table_sheet(ws, config, data)
    
    def _create_vertical_sheet(self, ws, config, data):
        """创建纵向布局工作表（如Vehicle Information）"""
        vehicle_info = data.get('vehicle', {}).get('vehicle_info', {})
        
        # 写入标题行
        headers = config['columns']
        for col_idx, col_config in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_config['header'])
            self._apply_style(cell, col_config.get('style', 'header'))
        
        # 按字段顺序写入数据
        row = 2
        for field_def in config.get('fields', []):
            field_key = field_def['field']
            display_name = field_def['display_name']
            value = vehicle_info.get(field_key, '')
            
            # 写入参数名
            cell = ws.cell(row=row, column=1, value=display_name)
            self._apply_style(cell, 'header_gray')
            
            # 写入值
            cell = ws.cell(row=row, column=2, value=value)
            self._apply_style(cell, 'cell_normal')
            
            row += 1
        
        # 设置列宽
        for col_config in config['columns']:
            col_idx = 1 if col_config['id'] == 'parameter' else 2
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_config['width']
    
    def _create_table_sheet(self, ws, config, data):
        """创建表格布局工作表"""
        columns = config['columns']
        
        # 写入表头
        for col_idx, col_config in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_config['header'])
            self._apply_style(cell, col_config.get('style', 'header'))
        
        # 写入数据
        self._write_table_data(ws, config, data, start_row=2)
        
        # 设置列宽
        for col_idx, col_config in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_config['width']
    
    def _write_table_data(self, ws, config, data, start_row=2):
        """写入表格数据"""
        sheet_name = config['name']
        row = start_row
        
        if sheet_name == 'Component Summary':
            self._write_component_summary(ws, config, data, start_row)
        elif sheet_name == 'Detailed Results':
            self._write_detailed_results(ws, config, data, start_row)
    
    def _write_component_summary(self, ws, config, data, start_row):
        """写入组件汇总数据"""
        components = data.get('components', {})
        row = start_row
        
        for comp_id, comp_data in components.items():
            values = {
                'component_id': comp_id,
                'channel_name': comp_data.get('channel_name', ''),
                'condition_count': len(comp_data.get('conditions', {})),
                'matched_count': sum(1 for c in comp_data.get('conditions', {}).values() 
                                    if c.get('matched', False)),
                'match_rate': 0.0  # 需要计算
            }
            
            if values['condition_count'] > 0:
                values['match_rate'] = values['matched_count'] / values['condition_count']
            
            for col_idx, col_config in enumerate(config['columns'], 1):
                field = col_config['field']
                value = values.get(field, '')
                cell = ws.cell(row=row, column=col_idx, value=value)
                self._apply_style(cell, col_config.get('style', 'cell_normal'))
            
            row += 1
    
    def _write_detailed_results(self, ws, config, data, start_row):
        """写入详细结果数据"""
        components = data.get('components', {})
        row = start_row
        seq = 1
        
        for comp_id, comp_data in components.items():
            for cond_id, cond_data in comp_data.get('conditions', {}).items():
                for col_idx, col_config in enumerate(config['columns'], 1):
                    field = col_config['field']
                    value = self._get_nested_value(cond_data, field, seq, comp_id, cond_id)
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    self._apply_style(cell, col_config.get('style', 'cell_normal'))
                
                row += 1
                seq += 1
    
    def _get_nested_value(self, data, field, seq, comp_id, cond_id):
        """获取嵌套字段值"""
        if field == 'sequence':
            return seq
        elif field == 'component_id':
            return comp_id
        elif field == 'condition_id':
            return cond_id
        elif field.startswith('statistics.'):
            stat_field = field.split('.')[1]
            return data.get('statistics', {}).get(stat_field, '')
        else:
            return data.get(field, '')
    
    def _apply_style(self, cell, style_name):
        """应用样式"""
        style_def = self.styles.get('styles', {}).get(style_name, {})
        
        # 应用字体
        if 'font' in style_def:
            font_config = self.styles.get('fonts', {}).get(style_def['font'], {})
            cell.font = Font(**font_config)
        
        # 应用填充
        if 'fill' in style_def:
            fill_config = self.styles.get('fills', {}).get(style_def['fill'], {})
            if fill_config.get('patternType') == 'solid':
                cell.fill = PatternFill(
                    patternType='solid',
                    fgColor=fill_config.get('fgColor', 'FFFFFF')
                )
        
        # 应用对齐
        if 'alignment' in style_def:
            align_config = self.styles.get('alignments', {}).get(style_def['alignment'], {})
            cell.alignment = Alignment(**align_config)
        
        # 应用边框
        if 'border' in style_def:
            border_config = self.styles.get('borders', {}).get(style_def['border'], {})
            cell.border = Border(**{
                side: Side(style=border_config[side]['style'], color=border_config[side]['color'])
                for side in ['left', 'right', 'top', 'bottom']
            })


def main():
    parser = argparse.ArgumentParser(description='Generate vehicle ripple test Excel report (V3.0)')
    parser.add_argument('--input-json', help='Input JSON file path')
    parser.add_argument('--vehicle-folder', help='Vehicle folder path')
    parser.add_argument('--output-excel', required=True, help='Output Excel file path')
    
    args = parser.parse_args()
    
    if args.input_json:
        # Load from JSON
        data = load_json_data(args.input_json)
    elif args.vehicle_folder:
        # Look for JSON in vehicle folder
        json_path = os.path.join(args.vehicle_folder, 'output.json')
        if os.path.exists(json_path):
            data = load_json_data(json_path)
        else:
            raise FileNotFoundError(f"output.json not found in {args.vehicle_folder}")
    else:
        raise ValueError("Please provide --input-json or --vehicle-folder")
    
    # Generate Excel report
    generate_excel_report(data, args.output_excel)

if __name__ == '__main__':
    main()
