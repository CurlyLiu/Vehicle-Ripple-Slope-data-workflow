#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vehicle Information Formatter - Unified formatting for ripple and slope skills

确保纹波技能和斜率技能生成的 Vehicle Information 完全一致:
1. 相同的字段顺序
2. 相同的字段显示名称
3. 相同的字段选择逻辑
4. JSON/DB/Excel 中的 vehicle_info 一致
"""

from typing import Dict, Any, List, Tuple, Optional


# 标准字段定义（顺序重要）
STANDARD_FIELDS = [
    ('vehicle_id', 'Vehicle ID'),
    ('vehicle_model', 'Vehicle Model'),
    ('manufacturer', 'Manufacturer'),
    ('length_mm', 'Length (mm)'),
    ('width_mm', 'Width (mm)'),
    ('height_mm', 'Height (mm)'),
    ('wheelbase_mm', 'Wheelbase (mm)'),
    ('front_track_mm', 'Front Track (mm)'),
    ('rear_track_mm', 'Rear Track (mm)'),
    ('min_ground_clearance_mm', 'Min Ground Clearance (mm)'),
]

# 需要排除的字段（不在Excel中显示）
EXCLUDED_FIELDS = {
    # 已标准化的字段
    'vehicle_id', 'vehicle_model', 'manufacturer',
    'length_mm', 'width_mm', 'height_mm',
    'wheelbase_mm', 'front_track_mm', 'rear_track_mm',
    'min_ground_clearance_mm',
    # 中文别名
    '车型', '制造商', '车长', '车宽', '车高', '车长mm', '车宽mm', '车高mm',
    '长度(mm)', '宽度(mm)', '高度(mm)', '轴距', '轴距(mm)',
    # 无效/分隔符字段
    '---', '参数名称', '',
}


def format_vehicle_info_for_excel(vehicle_data: Dict) -> List[Tuple[str, Any]]:
    """
    统一格式化 vehicle_info 为 Excel 显示格式

    Args:
        vehicle_data: 包含 vehicle 和 vehicle_info 的字典

    Returns:
        列表，每项为 (显示名称, 值) 元组
    """
    vehicle = vehicle_data.get('vehicle', {})
    vehicle_info = vehicle_data.get('vehicle_info', {})

    result = []

    # 1. 添加标准字段（按固定顺序）
    result.append(('Vehicle ID', vehicle.get('vehicle_id', 'Unknown')))

    for field_key, display_name in STANDARD_FIELDS[1:]:  # 跳过 vehicle_id
        value = vehicle_info.get(field_key)
        if value is not None and str(value).strip():
            result.append((display_name, value))

    # 2. 添加其他字段（按原始顺序，但排除已显示和排除列表中的字段）
    displayed_keys = {f[0] for f in STANDARD_FIELDS}

    for key, value in vehicle_info.items():
        # 跳过已显示的标准字段
        if key in displayed_keys:
            continue
        # 跳过排除列表中的字段
        if key in EXCLUDED_FIELDS:
            continue
        # 跳过空值
        if value is None or str(value).strip() == '':
            continue

        # 转换显示名称：如果有下划线转空格并首字母大写，否则保持原样
        if '_' in key and not any(c in key for c in '()（）'):
            display_key = key.replace('_', ' ').title()
        else:
            display_key = key

        result.append((display_key, value))

    return result


def format_vehicle_info_for_json(vehicle_data: Dict) -> Dict:
    """
    统一格式化 vehicle_info 为 JSON 存储格式

    确保 JSON 中的字段顺序和内容是确定的

    Args:
        vehicle_data: 包含 vehicle 和 vehicle_info 的字典

    Returns:
        标准化的 vehicle_info 字典
    """
    vehicle = vehicle_data.get('vehicle', {})
    vehicle_info = vehicle_data.get('vehicle_info', {})

    result = {
        'vehicle_id': vehicle.get('vehicle_id', 'Unknown'),
    }

    # 按标准顺序添加字段
    for field_key, _ in STANDARD_FIELDS[1:]:
        value = vehicle_info.get(field_key)
        if value is not None:
            result[field_key] = value

    # 添加其他字段（保持原始顺序）
    for key, value in vehicle_info.items():
        if key not in result and key not in EXCLUDED_FIELDS:
            if value is not None:
                result[key] = value

    return result


def create_vehicle_info_sheet_unified(wb, vehicle_data: Dict, skill_type: str = 'ripple'):
    """
    统一创建 Vehicle Information Excel 工作表

    Args:
        wb: openpyxl Workbook 对象
        vehicle_data: 车辆数据字典
        skill_type: 'ripple' 或 'slope'（用于输出文件夹命名）
    """
    from openpyxl.styles import Font, Alignment, PatternFill

    ws = wb.active
    ws.title = "Vehicle Information"

    # 添加标题行
    header_fill = PatternFill(start_color='D5E8F0', end_color='D5E8F0', fill_type='solid')

    cell = ws.cell(row=1, column=1, value='Parameter')
    cell.font = Font(bold=True, color='000000')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = ws.cell(row=1, column=2, value='Value')
    cell.font = Font(bold=True, color='000000')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # 获取统一格式化的数据
    fields = format_vehicle_info_for_excel(vehicle_data)

    # 填充数据（从第2行开始）
    for row_idx, (key, value) in enumerate(fields, start=2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=value)

        # 左对齐
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center')

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30

    return ws


def normalize_vehicle_info(data: Dict) -> Dict:
    """
    完全标准化 vehicle_info，确保两个技能输出完全一致

    用于在生成 JSON/DB 之前对数据进行规范化

    Args:
        data: 原始处理结果字典

    Returns:
        标准化后的数据字典
    """
    result = {
        'vehicle': {
            'vehicle_id': data.get('vehicle', {}).get('vehicle_id', 'Unknown'),
            'vehicle_info': {}
        },
        'components': data.get('components', {}),
        'metadata': data.get('metadata', {})
    }

    # 标准化 vehicle_info
    vehicle_info = data.get('vehicle', {}).get('vehicle_info', {})

    # 按标准顺序添加
    for field_key, _ in STANDARD_FIELDS[1:]:
        value = vehicle_info.get(field_key)
        if value is not None and str(value).strip():
            result['vehicle']['vehicle_info'][field_key] = value

    # 添加其他非排除字段
    for key, value in vehicle_info.items():
        if key not in EXCLUDED_FIELDS and key not in result['vehicle']['vehicle_info']:
            if value is not None and str(value).strip():
                result['vehicle']['vehicle_info'][key] = value

    return result


# 向后兼容的别名
format_vehicle_info = format_vehicle_info_for_excel
create_vehicle_info_sheet = create_vehicle_info_sheet_unified
