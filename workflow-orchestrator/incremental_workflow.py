"""
全工作流增量处理引擎

为每个阶段的输入计算指纹，与缓存对比判定是否需要重新执行。
支持: 单车辆增量、批量增量、强制全量重跑
"""

import json
import hashlib
import os
import re
import sqlite3
import sys
from pathlib import Path
from datetime import datetime, timezone
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, asdict


# 主进程输出无缓冲: 即使被管道(| tee)、文件重定向(> log)捕获,
# print() 也按行立即刷新,避免日志看起来"卡住不动"
try:
    sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)
    sys.stderr.reconfigure(encoding='utf-8', line_buffering=True)
except (AttributeError, ValueError):
    # Python <3.7 或某些环境(如 IDLE)不支持 reconfigure
    pass


@dataclass
class StageDecision:
    """阶段执行决策"""
    stage: str
    should_run: bool
    reason: str
    cached_fp: Optional[str]
    current_fp: Optional[str]
    estimated_time: int


class IncrementalWorkflow:
    """
    全工作流增量处理引擎

    缓存文件: {base_dir}/{vehicle_id}/.workflow_cache.json

    使用示例:
        workflow = IncrementalWorkflow("V0001", "F:/Vehicle_Date")
        plan = workflow.build_execution_plan()
        for decision in plan:
            if decision.should_run:
                workflow.execute_stage(decision.stage)
            else:
                print(f"跳过 {decision.stage}: {decision.reason}")
    """

    CACHE_FILENAME = ".workflow_cache.json"
    CACHE_SCHEMA_VERSION = 2  # P1.6 v1.4: schema 版本号,用于检测升级

    # C1 v1.4 修订: single-flight 改 class-level,batch_run 起始时重置
    # 之前是实例变量,但 batch_run 每辆车都 new IncrementalWorkflow → flag 重置
    # → 24 辆车都触发 stage4 重 import (而非只第一辆)
    _stage4_missing_handled_global = False

    @classmethod
    def reset_stage4_single_flight(cls):
        """C1 v1.4: batch_run 起始时调用,重置 stage4 single-flight 状态."""
        cls._stage4_missing_handled_global = False

    def __init__(self, vehicle_id: str, base_dir: str,
                 skills_dir: str = "C:/Users/31915/.claude/skills"):
        self.vehicle_id = vehicle_id
        self.base_dir = Path(base_dir)
        self.skills_dir = Path(skills_dir)
        self.vehicle_dir = self.base_dir / vehicle_id
        self.cache_path = self.vehicle_dir / self.CACHE_FILENAME
        self.cache = self._load_cache()
        self.decisions: List[StageDecision] = []
        self.execution_log: List[Dict] = []

    def _save_cache(self):
        """原子写 cache + .bak 回退 (CR-N8 v1.4 修订,C2 v1.4 修订加 try/finally + retry).

        旧实现 open('w') 直写,中途崩溃会损坏 cache 文件,下次 _load_cache
        捕获 JSONDecodeError -> return {} -> 视为新车辆 -> 全 stage 重跑。

        新实现:
        1. 写到 .tmp 临时文件 + fsync 刷盘
        2. 备份现有 cache 到 .bak (若存在)
        3. 原子 rename (os.replace 在 Windows 上是原子的) + retry 处理文件锁

        _load_cache 增加 .bak 回退路径,损坏时尝试从备份恢复。
        """
        import shutil
        import time
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        # P1.6 v1.4: 保存时标记 schema 版本
        self.cache['_schema_version'] = self.CACHE_SCHEMA_VERSION

        tmp = self.cache_path.with_suffix('.json.tmp')
        bak = self.cache_path.with_suffix('.json.bak')

        try:
            # 1. 写到 tmp + 强制刷盘
            with open(tmp, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, ensure_ascii=False, indent=2)
                f.flush()
                try:
                    os.fsync(f.fileno())  # 强制刷盘
                except (OSError, AttributeError):
                    pass  # 某些文件系统不支持 fsync,容忍

            # 2. 备份现有 cache (若存在,且非首次)
            if self.cache_path.exists():
                try:
                    shutil.copy2(self.cache_path, bak)
                except OSError:
                    pass  # 备份失败不阻断主流程

            # 3. 原子 rename (Windows: os.replace 是原子;但被其他进程打开时会 PermissionError)
            # C2 v1.4: retry 3 次,间隔 0.1s 缓解 Windows 文件锁
            last_error = None
            for attempt in range(3):
                try:
                    os.replace(tmp, self.cache_path)
                    last_error = None
                    break
                except PermissionError as e:
                    last_error = e
                    if attempt < 2:
                        time.sleep(0.1)
            if last_error is not None:
                raise last_error
        finally:
            # C2 v1.4: 确保 tmp 文件不留残
            if tmp.exists():
                try:
                    tmp.unlink()
                except OSError:
                    pass

    def _load_cache(self) -> Dict:
        if self.cache_path.exists():
            try:
                with open(self.cache_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # P1.6 v1.4: 检测 schema 升级
                if isinstance(data, dict):
                    schema = data.get('_schema_version', 1)
                    if schema < self.CACHE_SCHEMA_VERSION:
                        print(f"[INFO] {self.vehicle_id} cache schema v{schema} -> v{self.CACHE_SCHEMA_VERSION} 升级检测")
                        print(f"[INFO] 算法变更: stage2 现已纳入 vehicle_info.{{md,xlsx}} 指纹 (CR-N2)")
                        print(f"[INFO] 首次 batch 将触发 stage2 一次性重跑 (这是预期行为)")
                return data
            except (json.JSONDecodeError, UnicodeDecodeError) as e:
                # P1.9 v1.4: 损坏时尝试从 .bak 恢复
                bak = self.cache_path.with_suffix('.json.bak')
                if bak.exists():
                    print(f"[WARN] {self.vehicle_id} cache 损坏 ({type(e).__name__}: {e}),尝试从 .bak 恢复")
                    try:
                        with open(bak, 'r', encoding='utf-8') as f:
                            return json.load(f)
                    except Exception:
                        pass
                print(f"[ERROR] {self.vehicle_id} cache 完全损坏且无可用备份,视为新车辆 (将触发全 stage 重跑)")
                return {}
        return {}

    def clear_cache(self):
        """清空缓存,下次执行将全量重跑"""
        self.cache = {}
        # L1 v1.4 修订: 同时清理 .bak 和 .tmp,避免恢复时读到旧数据
        for suffix in ('.json', '.json.bak', '.json.tmp'):
            p = self.cache_path.with_suffix(suffix)
            if p.exists():
                try:
                    p.unlink()
                except OSError:
                    pass

    def _file_fingerprint(self, path: Path, algorithm: str = "fast") -> str:
        """
        计算文件指纹

        algorithm='fast': mtime + size (适合大文件，如.dmd)
        algorithm='sha256': 完整内容哈希 (适合小文件，如.xlsx/.md)
        """
        if not path.exists():
            return "MISSING"

        stat = path.stat()

        if algorithm == "fast":
            return f"{int(stat.st_mtime)}:{stat.st_size}"

        h = hashlib.sha256()
        with open(path, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b''):
                h.update(chunk)
        return h.hexdigest()[:16]

    def _dir_fingerprint(self, path: Path, pattern: str = "*",
                         algorithm: str = "fast") -> str:
        """计算目录指纹"""
        if not path.exists():
            return "MISSING"

        files = sorted(path.rglob(pattern) if '**' in pattern else path.glob(pattern))
        fingerprints = []

        for f in files:
            if f.is_file():
                fp = self._file_fingerprint(f, algorithm)
                fingerprints.append(f"{f.relative_to(path)}:{fp}")

        return hashlib.sha256("\n".join(fingerprints).encode()).hexdigest()[:16]

    def _files_fingerprint(self, paths: List[Path], algorithm: str = "sha256") -> str:
        """计算多个文件的聚合指纹 (v1.6 hotfix P3.1: 用 relative_to 而非 p.name 避免不同目录同名文件冲突)"""
        fingerprints = []
        for p in sorted(paths):
            fp = self._file_fingerprint(p, algorithm)
            try:
                rel = p.relative_to(self.vehicle_dir).as_posix()
            except ValueError:
                rel = str(p)  # 超出 vehicle_dir 时用绝对路径
            fingerprints.append(f"{rel}:{fp}")
        return hashlib.sha256("\n".join(fingerprints).encode()).hexdigest()[:16]

    def _semantic_fingerprint(self, path: Path) -> str:
        """语义指纹 (CR-N2 v1.4): 屏蔽非内容差异,避免误报变更.

        - .xlsx: 用 openpyxl 读 cells 后 hash,屏蔽 zip metadata/mtime 差异
        - .md/.txt: 规范化换行 (\\r\\n -> \\n) + 去尾空白,屏蔽编辑器换行差异
        - 其他: 回退到 sha256 _file_fingerprint

        若读取失败,回退到 _file_fingerprint 以保证不抛异常。
        """
        if not path.exists():
            return "MISSING"
        try:
            if path.suffix.lower() == '.xlsx':
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True, data_only=True)
                content = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    content.append(f"=={sheet}==")
                    for row in ws.iter_rows(values_only=True):
                        content.append("|".join("" if c is None else str(c) for c in row))
                wb.close()
                return hashlib.sha256("\n".join(content).encode('utf-8', errors='replace')).hexdigest()[:16]
            elif path.suffix.lower() in ('.md', '.txt'):
                # 规范化换行 + 去尾空白
                for enc in ('utf-8', 'gbk'):
                    try:
                        text = path.read_text(encoding=enc)
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    return self._file_fingerprint(path, 'sha256')
                normalized = text.replace('\r\n', '\n').replace('\r', '\n').rstrip()
                return hashlib.sha256(normalized.encode('utf-8')).hexdigest()[:16]
        except Exception:
            pass  # 任何读取失败回退到 sha256
        return self._file_fingerprint(path, 'sha256')

    def _files_fingerprint_smart(self, paths: List[Path]) -> str:
        """智能聚合指纹 (CR-N2 v1.4 兼容): vehicle_info.{md,xlsx} 走语义指纹,其他用 sha256.

        用于 stage2 输入指纹计算,避免 vehicle_info 文件因 mtime/换行变化误触重跑。
        v1.6 hotfix P3.1: 用 relative_to 而非 p.name 避免不同目录同名文件冲突。
        """
        fingerprints = []
        for p in sorted(paths):
            if p.name in ('vehicle_info.md', 'vehicle_info.xlsx'):
                fp = self._semantic_fingerprint(p)
            else:
                fp = self._file_fingerprint(p, 'sha256')
            try:
                rel = p.relative_to(self.vehicle_dir).as_posix()
            except ValueError:
                rel = str(p)
            fingerprints.append(f"{rel}:{fp}")
        return hashlib.sha256("\n".join(fingerprints).encode()).hexdigest()[:16]

    def _resolve_db_dir(self) -> Path:
        """从 ~/.vehicle_database/config.json 解析 DB 目录 (CR-2 v1.2 修订).

        优先级: 顶级 database_path > database.default_path > 默认 F:/Vehicle_Database
        若指向 .db 文件,自动提取 parent 目录。
        """
        from pathlib import Path as _P
        cfg_path = _P.home() / ".vehicle_database" / "config.json"
        db_dir = _P("F:/Vehicle_Database")  # 最终 fallback
        if cfg_path.exists():
            try:
                with open(cfg_path, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                # 优先级: 顶级 database_path > database.default_path
                raw = cfg.get("database_path") or cfg.get("database", {}).get("default_path")
                if raw:
                    p = _P(raw)
                    db_dir = p.parent if p.suffix == ".db" else p
            except Exception:
                pass  # 配置读取失败用默认
        return db_dir

    def _output_paths(self, stage: str) -> List[Path]:
        """返回指定 stage 的输出文件路径列表 (用于存在性检查).

        P1.4a v1.4 修订: 仅新增,不替换已有逻辑。stage4 路径从 config.json 解析。
        """
        root = self.vehicle_dir
        vid = self.vehicle_id
        if stage == "stage2_ripple":
            out = root / f"{vid}_RIPPLE" / f"{vid}_RIPPLE_output"
            return [out / f"{vid}_RIPPLE_data.json",
                    out / f"{vid}_RIPPLE.db",
                    out / f"{vid}_RIPPLE_summary.xlsx"]
        if stage == "stage2_slope":
            out = root / f"{vid}_SLOPE" / f"{vid}_SLOPE_output"
            return [out / f"{vid}_SLOPE_data.json",
                    out / f"{vid}_SLOPE.db",
                    out / f"{vid}_SLOPE_summary.xlsx"]
        if stage.startswith("stage3_ripple_"):
            comp = stage.replace("stage3_ripple_", "")
            return [root / f"{vid}_RIPPLE" / f"{vid}_RIPPLE_output"
                         / f"{vid}_RIPPLE_REPORT_{comp}.docx"]
        if stage.startswith("stage3_slope_"):
            comp = stage.replace("stage3_slope_", "")
            return [root / f"{vid}_SLOPE" / f"{vid}_SLOPE_output"
                         / f"{vid}_SLOPE_REPORT_{comp}.docx"]
        if stage == "stage4":
            db_dir = self._resolve_db_dir()
            return [db_dir / "Ripple.db", db_dir / "Slope.db"]
        raise ValueError(f"Unknown stage: {stage}")

    def _decide_stage1(self) -> StageDecision:
        """判定阶段1是否需要执行"""
        dmd_dir = self.vehicle_dir / "test_data"

        if not dmd_dir.exists():
            return StageDecision(
                stage="stage1", should_run=False,
                reason="无 test_data 目录", cached_fp=None, current_fp=None,
                estimated_time=0
            )

        current_fp = self._dir_fingerprint(dmd_dir, "*.dmd", algorithm="fast")
        cached = self.cache.get("stage1", {})
        cached_fp = cached.get("fingerprint")

        if cached_fp is None:
            return StageDecision(
                stage="stage1", should_run=True,
                reason="首次运行", cached_fp=None, current_fp=current_fp,
                estimated_time=3600
            )

        if current_fp != cached_fp:
            return StageDecision(
                stage="stage1", should_run=True,
                reason=f".dmd文件有变更 (指纹: {cached_fp[:8]} -> {current_fp[:8]})",
                cached_fp=cached_fp, current_fp=current_fp,
                estimated_time=3600
            )

        ripple_dir = self.vehicle_dir / f"{self.vehicle_id}_RIPPLE"
        slope_dir = self.vehicle_dir / f"{self.vehicle_id}_SLOPE"

        if not ripple_dir.exists() and not slope_dir.exists():
            return StageDecision(
                stage="stage1", should_run=True,
                reason=".dmd文件未变化，但无输出目录(可能上次被删除)",
                cached_fp=cached_fp, current_fp=current_fp,
                estimated_time=3600
            )

        return StageDecision(
            stage="stage1", should_run=False,
            reason=".dmd文件无变更，输出目录存在",
            cached_fp=cached_fp, current_fp=current_fp,
            estimated_time=0
        )

    def _decide_stage2_ripple(self) -> StageDecision:
        """判定阶段2(纹波)是否需要执行"""
        ripple_dir = self.vehicle_dir / f"{self.vehicle_id}_RIPPLE"

        if not ripple_dir.exists():
            return StageDecision(
                stage="stage2_ripple", should_run=False,
                reason="无 RIPPLE 数据目录", cached_fp=None, current_fp=None,
                estimated_time=0
            )

        inputs = []
        for xlsx in ripple_dir.rglob("statistics.xlsx"):
            inputs.append(xlsx)

        for rule_name in ["test_naming_rules.md", "sensor_naming_rules.md"]:
            rule_path = self.vehicle_dir / rule_name
            if rule_path.exists():
                inputs.append(rule_path)

        # P1.5 v1.4: vehicle_info 纳入指纹 (修改车型信息会触发重跑)
        for info_name in ["vehicle_info.md", "vehicle_info.xlsx"]:
            info_path = self.vehicle_dir / info_name
            if info_path.exists():
                inputs.append(info_path)

        current_fp = self._files_fingerprint_smart(inputs)
        cached = self.cache.get("stage2_ripple", {})
        cached_fp = cached.get("fingerprint")

        if cached_fp is None:
            return StageDecision(
                stage="stage2_ripple", should_run=True,
                reason="首次运行",
                cached_fp=None, current_fp=current_fp,
                estimated_time=600
            )

        if current_fp != cached_fp:
            return StageDecision(
                stage="stage2_ripple", should_run=True,
                reason=f"输入文件或规则有变更 (指纹: {cached_fp[:8]} -> {current_fp[:8]})",
                cached_fp=cached_fp, current_fp=current_fp,
                estimated_time=600
            )

        output_dir = ripple_dir / f"{self.vehicle_id}_RIPPLE_output"
        expected_files = [
            output_dir / f"{self.vehicle_id}_RIPPLE_data.json",
            output_dir / f"{self.vehicle_id}_RIPPLE.db",
            output_dir / f"{self.vehicle_id}_RIPPLE_summary.xlsx",
        ]

        missing = [f.name for f in expected_files if not f.exists()]
        if missing:
            return StageDecision(
                stage="stage2_ripple", should_run=True,
                reason=f"输出文件缺失: {', '.join(missing)}",
                cached_fp=cached_fp, current_fp=current_fp,
                estimated_time=600
            )

        return StageDecision(
            stage="stage2_ripple", should_run=False,
            reason="输入文件和规则无变更，输出文件完整",
            cached_fp=cached_fp, current_fp=current_fp,
            estimated_time=0
        )

    def _decide_stage2_slope(self) -> StageDecision:
        """判定阶段2(斜率)是否需要执行"""
        slope_dir = self.vehicle_dir / f"{self.vehicle_id}_SLOPE"

        if not slope_dir.exists():
            return StageDecision(
                stage="stage2_slope", should_run=False,
                reason="无 SLOPE 数据目录", cached_fp=None, current_fp=None,
                estimated_time=0
            )

        # 如果 RIPPLE 存在且会执行统一处理，则 SLOPE 由统一处理覆盖
        # vehicle_skills_cli.py process 会同时处理 RIPPLE 和 SLOPE
        ripple_dir = self.vehicle_dir / f"{self.vehicle_id}_RIPPLE"
        if ripple_dir.exists():
            ripple_decision = self._decide_stage2_ripple()
            if ripple_decision.should_run:
                return StageDecision(
                    stage="stage2_slope", should_run=False,
                    reason="由 stage2_ripple 统一处理",
                    cached_fp=None, current_fp=None,
                    estimated_time=0
                )

        inputs = []
        for xlsx in slope_dir.rglob("statistics.xlsx"):
            inputs.append(xlsx)

        for rule_name in ["test_naming_rules.md", "sensor_naming_rules.md"]:
            rule_path = self.vehicle_dir / rule_name
            if rule_path.exists():
                inputs.append(rule_path)

        # P1.5 v1.4: vehicle_info 纳入指纹
        for info_name in ["vehicle_info.md", "vehicle_info.xlsx"]:
            info_path = self.vehicle_dir / info_name
            if info_path.exists():
                inputs.append(info_path)

        current_fp = self._files_fingerprint_smart(inputs)
        cached = self.cache.get("stage2_slope", {})
        cached_fp = cached.get("fingerprint")

        if cached_fp is None:
            return StageDecision(
                stage="stage2_slope", should_run=True,
                reason="首次运行",
                cached_fp=None, current_fp=current_fp,
                estimated_time=600
            )

        if current_fp != cached_fp:
            return StageDecision(
                stage="stage2_slope", should_run=True,
                reason=f"输入文件或规则有变更 (指纹: {cached_fp[:8]} -> {current_fp[:8]})",
                cached_fp=cached_fp, current_fp=current_fp,
                estimated_time=600
            )

        output_dir = slope_dir / f"{self.vehicle_id}_SLOPE_output"
        expected_files = [
            output_dir / f"{self.vehicle_id}_SLOPE_data.json",
            output_dir / f"{self.vehicle_id}_SLOPE.db",
            output_dir / f"{self.vehicle_id}_SLOPE_summary.xlsx",
        ]

        missing = [f.name for f in expected_files if not f.exists()]
        if missing:
            return StageDecision(
                stage="stage2_slope", should_run=True,
                reason=f"输出文件缺失: {', '.join(missing)}",
                cached_fp=cached_fp, current_fp=current_fp,
                estimated_time=600
            )

        return StageDecision(
            stage="stage2_slope", should_run=False,
            reason="输入文件和规则无变更，输出文件完整",
            cached_fp=cached_fp, current_fp=current_fp,
            estimated_time=0
        )

    def _decide_stage4(self) -> StageDecision:
        """判定阶段4(数据库导入)是否需要执行"""
        ripple_json = self.vehicle_dir / f"{self.vehicle_id}_RIPPLE" / f"{self.vehicle_id}_RIPPLE_output" / f"{self.vehicle_id}_RIPPLE_data.json"
        slope_json = self.vehicle_dir / f"{self.vehicle_id}_SLOPE" / f"{self.vehicle_id}_SLOPE_output" / f"{self.vehicle_id}_SLOPE_data.json"

        inputs = [p for p in [ripple_json, slope_json] if p.exists()]

        if not inputs:
            return StageDecision(
                stage="stage4", should_run=False,
                reason="无阶段2输出可导入", cached_fp=None, current_fp=None,
                estimated_time=0
            )

        current_fp = self._files_fingerprint(inputs, algorithm="sha256")
        cached = self.cache.get("stage4", {})
        cached_fp = cached.get("fingerprint")

        if cached_fp != current_fp:
            return StageDecision(
                stage="stage4", should_run=True,
                reason="阶段2输出有变更" if cached_fp else "首次运行",
                cached_fp=cached_fp, current_fp=current_fp,
                estimated_time=180
            )

        # P1.4b v1.4 修订: 输入未变,检查全局 DB 存在性 (single-flight 守护)
        # 全局 Ripple.db/Slope.db 是 stage4 输出,如果被误删,需要触发重 import
        # C1 v1.4: 改用 class-level flag,batch_run 起始时重置
        if not type(self)._stage4_missing_handled_global:
            try:
                output_paths = self._output_paths("stage4")
                output_missing = [p for p in output_paths if not p.exists()]
                if output_missing:
                    type(self)._stage4_missing_handled_global = True
                    missing_names = [p.name for p in output_missing]
                    return StageDecision(
                        stage="stage4", should_run=True,
                        reason=f"全局数据库缺失: {missing_names}",
                        cached_fp=cached_fp, current_fp=current_fp,
                        estimated_time=180
                    )
            except Exception:
                pass  # 路径解析失败不影响主流程

        return StageDecision(
            stage="stage4", should_run=False,
            reason="阶段2输出无变更",
            cached_fp=cached_fp, current_fp=current_fp,
            estimated_time=0
        )

    def _decide_stage3_component(self, component_code: str, report_type: str) -> StageDecision:
        """判定阶段3某通道是否需要生成报告"""
        output_dir = self.vehicle_dir / f"{self.vehicle_id}_{report_type.upper()}"
        summary_file = output_dir / f"{self.vehicle_id}_{report_type.upper()}_output" / f"{self.vehicle_id}_{report_type.upper()}_summary.xlsx"

        template_path = self.skills_dir / "vehicle-report-generation" / "templates" / f"{report_type}_report_template.docx"

        inputs = [p for p in [summary_file, template_path] if p.exists()]

        if not summary_file.exists():
            return StageDecision(
                stage=f"stage3_{report_type}_{component_code}", should_run=False,
                reason="无汇总文件", cached_fp=None, current_fp=None,
                estimated_time=0
            )

        current_fp = self._files_fingerprint(inputs, algorithm="sha256")
        cached = self.cache.get(f"stage3_{report_type}_{component_code}", {})
        cached_fp = cached.get("fingerprint")

        report_path = output_dir / f"{self.vehicle_id}_{report_type.upper()}_output" / f"{self.vehicle_id}_{report_type.upper()}_REPORT_{component_code}.docx"

        if cached_fp != current_fp or not report_path.exists():
            reason = "数据或模板有变更" if cached_fp else "首次生成"
            if not report_path.exists():
                reason += ", 报告文件不存在"

            return StageDecision(
                stage=f"stage3_{report_type}_{component_code}", should_run=True,
                reason=reason, cached_fp=cached_fp, current_fp=current_fp,
                estimated_time=120
            )

        return StageDecision(
            stage=f"stage3_{report_type}_{component_code}", should_run=False,
            reason="数据和模板无变更，报告已存在",
            cached_fp=cached_fp, current_fp=current_fp,
            estimated_time=0
        )

    def build_execution_plan(self) -> List[StageDecision]:
        """构建完整执行计划"""
        plan = []

        plan.append(self._decide_stage1())
        plan.append(self._decide_stage2_ripple())
        plan.append(self._decide_stage2_slope())

        # 阶段3: 报告生成（逐个通道）
        components = self._get_components()
        for comp in components:
            plan.append(self._decide_stage3_component(comp, "ripple"))
            plan.append(self._decide_stage3_component(comp, "slope"))

        # 阶段4: 数据库导入
        plan.append(self._decide_stage4())

        self.decisions = plan
        return plan

    def _get_components(self) -> List[str]:
        """获取车辆的所有组件通道"""
        components = set()

        ripple_dir = self.vehicle_dir / f"{self.vehicle_id}_RIPPLE"
        if ripple_dir.exists():
            for subdir in ripple_dir.iterdir():
                if subdir.is_dir() and not subdir.name.endswith("_output"):
                    components.add(subdir.name)

        slope_dir = self.vehicle_dir / f"{self.vehicle_id}_SLOPE"
        if slope_dir.exists():
            for subdir in slope_dir.iterdir():
                if subdir.is_dir() and not subdir.name.endswith("_output"):
                    components.add(subdir.name)

        return sorted(components)

    def execute_stage(self, stage: str):
        """执行指定阶段"""
        import subprocess
        import sys

        start_time = datetime.now(timezone.utc)

        if stage == "stage1":
            print("阶段1 (AutoHandleFiles) 需手动执行 GUI，本引擎从阶段2开始增量处理")
            # C7 v1.6 hotfix: 提示中的车辆 ID 改用 self.vehicle_id,
            # 旧版硬编码 'V0001' 导致批量处理 V0002~V0024 时仍提示 run V0001,误导用户.
            print(f"   如已完成 GUI 操作，请直接运行: python incremental_workflow.py run {self.vehicle_id}")
            # NEW-3 v1.4: 改 status 为 "manual_required" 避免污染 cache
            # 之前用 status="success" 会触发 _update_stage_cache(stage1) 写入虚假完成标记
            result = {"status": "manual_required", "stage": stage, "note": "stage1 GUI not executed"}

        elif stage == "stage2_ripple":
            cmd = [
                sys.executable,
                str(self.skills_dir / "vehicle-ripple-data" / "scripts" / "cli" / "vehicle_skills_cli.py"),
                "process", str(self.vehicle_dir), "--progress", "--no-auto-db"
            ]
            result = self._run_command(cmd, stage)

        elif stage == "stage2_slope":
            cmd = [
                sys.executable,
                str(self.skills_dir / "vehicle-slope-data" / "scripts" / "cli" / "process_slope.py"),
                "process", "--folder", str(self.vehicle_dir / f"{self.vehicle_id}_SLOPE")
            ]
            result = self._run_command(cmd, stage)

        elif stage.startswith("stage3_"):
            parts = stage.split("_")
            report_type = parts[1]
            component = "_".join(parts[2:])

            cmd = [
                sys.executable,
                str(self.skills_dir / "vehicle-report-generation" / "vehicle_report_cli.py"),
                "generate", self.vehicle_id,
                "--type", report_type,
                "--component", component,
                "--base-dir", str(self.base_dir)
            ]
            result = self._run_command(cmd, stage)

        elif stage == "stage4":
            # 自动初始化数据库（若不存在）
            db_dir = self._resolve_db_dir()
            if not (db_dir / "Ripple.db").exists():
                init_cmd = [
                    sys.executable,
                    str(self.skills_dir / "vehicle-database" / "vehicle_database.py"),
                    "-s", str(self.base_dir),
                    "init", "-o", str(db_dir), "--yes"
                ]
                self._run_command(init_cmd, "stage4_init")
            cmd = [
                sys.executable,
                str(self.skills_dir / "vehicle-database" / "vehicle_database.py"),
                "-s", str(self.base_dir),
                "add", self.vehicle_id
            ]
            result = self._run_command(cmd, stage)

        else:
            result = {"status": "unknown_stage", "stage": stage}

        elapsed = (datetime.now(timezone.utc) - start_time).total_seconds()
        result["elapsed_sec"] = elapsed
        self.execution_log.append(result)

        if result.get("status") == "success":
            self._update_stage_cache(stage)

        return result

    def _run_command(self, cmd: List[str], stage: str) -> Dict:
        """执行子进程命令"""
        import subprocess

        # 子进程无缓冲: 让 vehicle_skills_cli.py 等子脚本的 print 实时输出,
        # 避免 capture_output 触发块缓冲导致日志延迟
        env = os.environ.copy()
        env['PYTHONUNBUFFERED'] = '1'
        env['PYTHONIOENCODING'] = 'utf-8'

        try:
            result = subprocess.run(
                cmd, capture_output=True, text=True,
                timeout=7200, encoding='utf-8', errors='replace',
                env=env,
            )

            # P1.8 + CR-N7 v1.4: 区分 success / partial / failed
            # add.py 新增 exit 3 = 部分失败 (success_count < len(vehicles))
            # partial 状态不更新 cache,下次 batch 重试
            rc = result.returncode
            if rc == 0:
                status = "success"
            elif rc == 3:
                status = "partial"
            else:
                status = "failed"

            return {
                "stage": stage,
                "status": status,
                "returncode": rc,
                "stdout": result.stdout[-2000:] if len(result.stdout) > 2000 else result.stdout,
                "stderr": result.stderr[-1000:] if len(result.stderr) > 1000 else result.stderr,
            }
        except subprocess.TimeoutExpired:
            return {"stage": stage, "status": "timeout", "error": "执行超时(>2小时)"}
        except Exception as e:
            return {"stage": stage, "status": "error", "error": str(e)}

    def _update_stage_cache(self, stage: str):
        """根据当前状态更新缓存"""
        if stage == "stage1":
            fp = self._dir_fingerprint(
                self.vehicle_dir / "test_data", "*.dmd", algorithm="fast"
            )
            self.cache["stage1"] = {
                "fingerprint": fp,
                "completed_at": datetime.now(timezone.utc).isoformat()
            }

        elif stage == "stage2_ripple":
            ripple_dir = self.vehicle_dir / f"{self.vehicle_id}_RIPPLE"
            inputs = list(ripple_dir.rglob("statistics.xlsx"))
            for rule_name in ["test_naming_rules.md", "sensor_naming_rules.md"]:
                rule_path = self.vehicle_dir / rule_name
                if rule_path.exists():
                    inputs.append(rule_path)
            # P1.5 v1.4: vehicle_info 纳入指纹 (与 _decide_stage2_ripple 同步)
            for info_name in ["vehicle_info.md", "vehicle_info.xlsx"]:
                info_path = self.vehicle_dir / info_name
                if info_path.exists():
                    inputs.append(info_path)
            fp = self._files_fingerprint_smart(inputs)
            self.cache["stage2_ripple"] = {
                "fingerprint": fp,
                "completed_at": datetime.now(timezone.utc).isoformat()
            }

        elif stage == "stage2_slope":
            slope_dir = self.vehicle_dir / f"{self.vehicle_id}_SLOPE"
            inputs = list(slope_dir.rglob("statistics.xlsx"))
            for rule_name in ["test_naming_rules.md", "sensor_naming_rules.md"]:
                rule_path = self.vehicle_dir / rule_name
                if rule_path.exists():
                    inputs.append(rule_path)
            # P1.5 v1.4: vehicle_info 纳入指纹 (与 _decide_stage2_slope 同步)
            for info_name in ["vehicle_info.md", "vehicle_info.xlsx"]:
                info_path = self.vehicle_dir / info_name
                if info_path.exists():
                    inputs.append(info_path)
            fp = self._files_fingerprint_smart(inputs)
            self.cache["stage2_slope"] = {
                "fingerprint": fp,
                "completed_at": datetime.now(timezone.utc).isoformat()
            }

        elif stage == "stage4":
            ripple_json = self.vehicle_dir / f"{self.vehicle_id}_RIPPLE" / f"{self.vehicle_id}_RIPPLE_output" / f"{self.vehicle_id}_RIPPLE_data.json"
            slope_json = self.vehicle_dir / f"{self.vehicle_id}_SLOPE" / f"{self.vehicle_id}_SLOPE_output" / f"{self.vehicle_id}_SLOPE_data.json"
            inputs = [p for p in [ripple_json, slope_json] if p.exists()]
            fp = self._files_fingerprint(inputs, algorithm="sha256")
            self.cache["stage4"] = {
                "fingerprint": fp,
                "completed_at": datetime.now(timezone.utc).isoformat()
            }

        elif stage.startswith("stage3_"):
            parts = stage.split("_")
            report_type = parts[1]
            component = "_".join(parts[2:])
            output_dir = self.vehicle_dir / f"{self.vehicle_id}_{report_type.upper()}"
            summary_file = output_dir / f"{self.vehicle_id}_{report_type.upper()}_output" / f"{self.vehicle_id}_{report_type.upper()}_summary.xlsx"
            template_path = self.skills_dir / "vehicle-report-generation" / "templates" / f"{report_type}_report_template.docx"
            inputs = [p for p in [summary_file, template_path] if p.exists()]
            fp = self._files_fingerprint(inputs, algorithm="sha256")
            self.cache[stage] = {
                "fingerprint": fp,
                "completed_at": datetime.now(timezone.utc).isoformat()
            }

        self._save_cache()

    def print_plan(self):
        """打印执行计划"""
        print(f"\n{'='*70}")
        print(f"车辆 {self.vehicle_id} 增量处理执行计划")
        print(f"{'='*70}")

        total_estimated = 0
        for d in self.decisions:
            status = "[执行]" if d.should_run else "[跳过]"
            print(f"{status} [{d.stage:30s}] {d.reason}")
            if d.should_run:
                total_estimated += d.estimated_time

        will_run = sum(1 for d in self.decisions if d.should_run)
        will_skip = sum(1 for d in self.decisions if not d.should_run)

        print(f"{'='*70}")
        print(f"总计: {will_run} 个阶段需执行, {will_skip} 个阶段可跳过")
        print(f"预估总耗时: {total_estimated//60} 分钟")

        # P3.5 + HR-N2 v1.4: 若有 stage3 需执行,显示 template fingerprint
        # 帮助用户识别"为何全员重跑 stage3" (template 改 1 字符即触发 24 辆 x N 通道 重跑)
        stage3_run = [d for d in self.decisions if d.should_run and d.stage.startswith("stage3_")]
        if stage3_run:
            for report_type in ("ripple", "slope"):
                if any(d.stage.startswith(f"stage3_{report_type}") for d in stage3_run):
                    tpl = self.skills_dir / "vehicle-report-generation" / "templates" / f"{report_type}_report_template.docx"
                    if tpl.exists():
                        stat = tpl.stat()
                        from datetime import datetime as _dt
                        mtime = _dt.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(timespec='seconds')
                        fp = self._file_fingerprint(tpl, 'sha256')
                        print(f"[INFO] {report_type} template: mtime={mtime} sha256={fp}")
        print(f"{'='*70}\n")

    def save_execution_report(self):
        """保存执行报告"""
        report_path = self.vehicle_dir / ".workflow_execution_log.json"

        report = {
            "vehicle_id": self.vehicle_id,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "plan": [asdict(d) for d in self.decisions],
            "execution": self.execution_log,
        }

        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)


def scan_vehicles(base_dir: Path) -> List[str]:
    """扫描目录下所有匹配 V\\d+ 模式的车辆文件夹，按字母排序返回"""
    vehicles = []
    for item in sorted(base_dir.iterdir()):
        # NEW-4 v1.4: 严格匹配 V<digits>$,避免匹配 V0001abc / V0001.backup 等
        if item.is_dir() and re.fullmatch(r'V\d+', item.name):
            vehicles.append(item.name)
    return vehicles


def batch_run(base_dir: Path, skills_dir: str, stages: str, force: bool):
    """批量执行增量工作流"""
    # C1 v1.4: batch 起始时重置 stage4 single-flight 状态
    IncrementalWorkflow.reset_stage4_single_flight()

    vehicles = scan_vehicles(base_dir)

    if not vehicles:
        print(f"未在 {base_dir} 下发现车辆文件夹")
        return

    print(f"\n{'='*70}")
    print(f"批量增量处理")
    print(f"{'='*70}")
    print(f"扫描到 {len(vehicles)} 辆车: {', '.join(vehicles)}\n")

    results = []
    total_start = datetime.now(timezone.utc)

    # P2.5 + P2.7 + CR-N9 v1.4: 增量原子写 batch_log 辅助函数
    def _save_batch_log_atomic(partial: bool = True):
        """每辆车完成后即写 batch_log,避免 batch 中途崩溃丢失全部结果."""
        current_total = (datetime.now(timezone.utc) - total_start).total_seconds()
        ok = sum(1 for r in results if r['status'] == 'OK')
        skip = sum(1 for r in results if r['status'] == 'SKIP')
        # H1 v1.4: PARTIAL 单独统计,不混入 fail
        partial_count = sum(1 for r in results if r['status'] == 'PARTIAL')
        fail = sum(1 for r in results if r['status'] in ('FAIL', 'ERROR'))
        batch_log_data = {
            "scanned_at": datetime.now(timezone.utc).isoformat(),
            "base_dir": str(base_dir),
            "total_vehicles": len(vehicles),
            "processed_so_far": len(results),
            "ok": ok, "skip": skip, "partial": partial_count, "fail": fail,
            "total_duration_sec": current_total,
            "force": force,
            "stages_filter": stages,
            "partial_log": partial,  # True = 增量写入中,False = 完整结束
            "results": results,
        }
        log_path = base_dir / ".workflow_batch_log.json"
        tmp = log_path.with_suffix(".json.tmp")
        try:
            with open(tmp, 'w', encoding='utf-8') as f:
                json.dump(batch_log_data, f, ensure_ascii=False, indent=2)
                f.flush()
                try:
                    os.fsync(f.fileno())
                except (OSError, AttributeError):
                    pass
            os.replace(tmp, log_path)
        except Exception as e:
            print(f"[WARN] batch_log 写入失败: {e}")

    for i, vid in enumerate(vehicles, 1):
        print(f"\n[{i}/{len(vehicles)}] 处理 {vid}...")
        print("-" * 70)

        # P2.8 + CR-N11 v1.4: 启动预检 - 警告非标准 vehicle_info* 文件
        # (如 V0007/vehicle_info1.xlsx 这种旧备份,被引擎完全忽略)
        vdir = base_dir / vid
        if vdir.exists():
            extras = [p for p in vdir.glob('vehicle_info*')
                     if p.name not in ('vehicle_info.md', 'vehicle_info.xlsx')]
            if extras:
                names = [p.name for p in extras]
                print(f"  [WARN] {vid}: 发现非标准 vehicle_info 文件 {names},被引擎忽略 (仅识别 vehicle_info.md/xlsx)")

        vehicle_start = datetime.now(timezone.utc)

        # P2.5 + CR-N9 v1.4: 单车 try/except 包裹,避免非 subprocess 异常中断整个 batch
        try:
            workflow = IncrementalWorkflow(vid, str(base_dir), skills_dir)

            if force:
                workflow.clear_cache()

            plan = workflow.build_execution_plan()

            if stages == "all":
                stages_to_run = [d for d in plan if d.should_run]
            else:
                target = stages
                stages_to_run = [d for d in plan if d.should_run and target in d.stage]

            # 判断阶段2的实际执行模式（基于 stages_to_run，而非 plan）
            stage2_ripple_executed = any(d.stage == "stage2_ripple" for d in stages_to_run)
            stage2_slope_executed = any(d.stage == "stage2_slope" for d in stages_to_run)
            if stage2_ripple_executed and not stage2_slope_executed:
                stage2_display = "执行(R+S)"
            elif stage2_ripple_executed and stage2_slope_executed:
                stage2_display = "执行(R+S)"
            elif not stage2_ripple_executed and stage2_slope_executed:
                stage2_display = "执行(S)"
            else:
                stage2_display = "跳过"

            # 判断阶段3的实际执行（报告生成）
            stage3_ripple_planned = sum(1 for d in stages_to_run if d.stage.startswith("stage3_ripple"))
            stage3_slope_planned = sum(1 for d in stages_to_run if d.stage.startswith("stage3_slope"))
            stage3_total = stage3_ripple_planned + stage3_slope_planned

            # 判断阶段4的实际执行（数据库导入）
            stage4_executed = any(d.stage == "stage4" for d in stages_to_run)
            stage4_display = "执行" if stage4_executed else "跳过"

            stage3_display = "跳过" if stage3_total == 0 else "执行(0/0)"

            if not stages_to_run:
                print(f"  所有阶段均为最新，无需处理")
                results.append({
                    'vehicle_id': vid,
                    'status': 'SKIP',
                    'stage2': stage2_display,
                    'stage3': stage3_display,
                    'stage4': stage4_display,
                    'error': None,
                    'duration': 0.0,
                })
                _save_batch_log_atomic(partial=True)
                continue

            # 执行阶段
            failed_stage = None
            partial_stages = []  # H1/H5 v1.4: 跟踪 partial 状态的 stage
            replanned_flag = False  # P1.4c: 二次规划只做一次,避免无限循环
            for decision in stages_to_run:
                print(f"  执行 [{decision.stage}]...")
                result = workflow.execute_stage(decision.stage)

                if result.get("status") == "success":
                    print(f"  [{decision.stage}] 完成 ({result.get('elapsed_sec', 0):.1f}s)")
                    # P1.4c v1.4: stage2 完成后二次规划,挂入新出现的 should_run 项
                    if (decision.stage in ("stage2_ripple", "stage2_slope")
                            and not replanned_flag):
                        new_plan = workflow.build_execution_plan()
                        existing_stages = {d.stage for d in stages_to_run}
                        added = []
                        for d in new_plan:
                            if d.should_run and d.stage not in existing_stages:
                                stages_to_run.append(d)
                                added.append(d.stage)
                        if added:
                            print(f"  [二次规划] stage2 完成后新增 {len(added)} 个阶段: {', '.join(added)}")
                        replanned_flag = True
                elif result.get("status") == "manual_required":
                    # NEW-3 v1.4: stage1 manual_required 不视为失败,不更新 cache,继续后续阶段
                    print(f"  [{decision.stage}] 需手动执行 GUI (跳过,cache 不更新)")
                elif result.get("status") == "partial":
                    # CR-N7 + P1.8 v1.4: 部分失败 (exit 3),不更新 cache,下次重试
                    # H1/H5 v1.4: 记录 partial,显示用,results 状态为 PARTIAL
                    print(f"  [{decision.stage}] 部分失败 (exit 3,cache 不更新,下次重试)")
                    partial_stages.append(decision.stage)
                    if decision.stage == "stage4":
                        IncrementalWorkflow.reset_stage4_single_flight()
                        stage4_display = "部分失败"  # H5
                else:
                    print(f"  [{decision.stage}] 失败")
                    failed_stage = decision.stage
                    if decision.stage == "stage4":
                        IncrementalWorkflow.reset_stage4_single_flight()
                    break

            workflow.save_execution_report()

            # 统计阶段3成功数（报告生成）
            stage3_ripple_ok = sum(1 for log in workflow.execution_log
                                   if log.get('stage', '').startswith("stage3_ripple") and log.get('status') == 'success')
            stage3_slope_ok = sum(1 for log in workflow.execution_log
                                  if log.get('stage', '').startswith("stage3_slope") and log.get('status') == 'success')
            stage3_ok = stage3_ripple_ok + stage3_slope_ok

            if stage3_total > 0:
                stage3_display = f"执行({stage3_ok}/{stage3_total})"
            else:
                stage3_display = "跳过"

            duration = (datetime.now(timezone.utc) - vehicle_start).total_seconds()

            if failed_stage:
                results.append({
                    'vehicle_id': vid,
                    'status': 'FAIL',
                    'stage2': stage2_display,
                    'stage3': stage3_display,
                    'stage4': stage4_display,
                    'error': f"{failed_stage} 失败",
                    'duration': duration,
                })
            elif partial_stages:
                # H1 v1.4: partial 状态全链路传播,不要错误标 OK
                results.append({
                    'vehicle_id': vid,
                    'status': 'PARTIAL',
                    'stage2': stage2_display,
                    'stage3': stage3_display,
                    'stage4': stage4_display,
                    'error': f"部分失败 stages: {', '.join(partial_stages)}",
                    'duration': duration,
                })
            else:
                results.append({
                    'vehicle_id': vid,
                    'status': 'OK',
                    'stage2': stage2_display,
                    'stage3': stage3_display,
                    'stage4': stage4_display,
                    'error': None,
                    'duration': duration,
                })

        except Exception as e:
            # CR-N9 v1.4: 单车异常 (磁盘满/权限/KeyError 等),记录但不中断 batch
            duration = (datetime.now(timezone.utc) - vehicle_start).total_seconds()
            err_msg = f"{type(e).__name__}: {e}"
            print(f"  [ERROR] {vid} 整车异常: {err_msg}")
            results.append({
                'vehicle_id': vid,
                'status': 'ERROR',
                'stage2': '?',
                'stage3': '?',
                'stage4': '?',
                'error': err_msg,
                'duration': duration,
            })
        finally:
            # P2.7 + CR-N9 v1.4: 每辆车完成后增量写 batch_log (原子写)
            _save_batch_log_atomic(partial=True)

    total_duration = (datetime.now(timezone.utc) - total_start).total_seconds()

    # 汇总报告 (H1 v1.4: PARTIAL 单独统计,不混入 fail)
    ok_count = sum(1 for r in results if r['status'] == 'OK')
    skip_count = sum(1 for r in results if r['status'] == 'SKIP')
    partial_count = sum(1 for r in results if r['status'] == 'PARTIAL')
    fail_count = sum(1 for r in results if r['status'] in ('FAIL', 'ERROR'))

    print(f"\n{'='*70}")
    print(f"批量增量处理汇总")
    print(f"{'='*70}")
    print(f"总车辆数: {len(vehicles)}")
    print(f"成功: {ok_count}")
    print(f"无需处理: {skip_count}")
    if partial_count > 0:
        print(f"部分失败 (PARTIAL): {partial_count}")
    print(f"失败: {fail_count}")
    print(f"总耗时: {total_duration:.1f}s")

    if fail_count > 0:
        failed_ids = [r['vehicle_id'] for r in results if r['status'] in ('FAIL', 'ERROR')]
        print(f"失败车辆: {', '.join(failed_ids)}")

    print(f"\n{'Vehicle ID':<12} {'阶段2':<12} {'阶段3':<8} {'阶段4':<12} {'状态':<8} {'耗时':<8}")
    print("-" * 70)
    for r in results:
        print(f"{r['vehicle_id']:<12} {r['stage2']:<12} {r['stage3']:<8} {r['stage4']:<12} {r['status']:<8} {r['duration']:<8.1f}")
    print(f"{'='*70}")

    # P2.7 + CR-N9 v1.4: 最终原子写 batch_log,标记 partial=False
    _save_batch_log_atomic(partial=False)
    log_path = base_dir / ".workflow_batch_log.json"
    print(f"\n批量日志已保存: {log_path}")


def main():
    import argparse
    import sys
    import shutil

    parser = argparse.ArgumentParser(description="全工作流增量处理引擎")
    parser.add_argument("command", choices=["plan", "run", "clear-cache", "batch", "snapshot"],
                        help="plan=仅生成计划, run=执行, clear-cache=清空缓存, batch=批量处理, snapshot=跨平台快照打包")
    parser.add_argument("vehicle_id", nargs="?", default=None, help="车辆ID (plan/run/clear-cache 需要)")
    parser.add_argument("--base-dir", default="F:/Vehicle_Date",
                        help="车辆数据根目录")
    parser.add_argument("--skills-dir", default="C:/Users/31915/.claude/skills",
                        help="技能安装目录")
    parser.add_argument("--force", action="store_true",
                        help="强制全量重跑(忽略缓存)")
    parser.add_argument("--stages", default="all",
                        help="指定阶段: all, 1, 2, 3, 4, 2_ripple, 2_slope")
    parser.add_argument("--scan", type=Path, default=None,
                        help="批量扫描目录 (batch 命令需要)")
    parser.add_argument("--source", type=Path, default=None,
                        help="snapshot 源目录 (snapshot 命令需要)")
    parser.add_argument("--output", type=Path, default=None,
                        help="snapshot 输出 .zip 路径 (snapshot 命令需要)")

    args = parser.parse_args()

    if args.command == "snapshot":
        # v1.6 hotfix P1.2: CR-N6 跨平台快照命令
        # Python shutil.make_archive 跨平台 (Win/Linux/macOS),
        # 替代 tar/zip 等平台依赖命令,符合 CR-N6 要求
        if not args.source or not args.output:
            print("错误: snapshot 需要 --source 和 --output")
            sys.exit(1)
        source_dir = args.source.resolve()
        output_path = args.output.resolve()
        if not source_dir.exists():
            print(f"[ERROR] Source not found: {source_dir}")
            sys.exit(1)
        # make_archive 需要不带后缀的 base name
        base_name = str(output_path.with_suffix(''))
        print(f"Creating snapshot: {source_dir} -> {output_path}")
        archive = shutil.make_archive(base_name, 'zip', root_dir=str(source_dir))
        size_mb = Path(archive).stat().st_size / 1024 / 1024
        print(f"Snapshot created: {archive} ({size_mb:.2f} MB)")
        return

    if args.command == "batch":
        scan_dir = args.scan if args.scan else Path(args.base_dir)
        batch_run(scan_dir, args.skills_dir, args.stages, args.force)
        return

    if not args.vehicle_id:
        print("错误: plan/run/clear-cache 命令需要提供 vehicle_id")
        sys.exit(1)

    workflow = IncrementalWorkflow(
        args.vehicle_id, args.base_dir, args.skills_dir
    )

    if args.force:
        workflow.clear_cache()
        print("缓存已清空，将执行全量处理\n")

    if args.command == "clear-cache":
        workflow.clear_cache()
        print("缓存已清空")
        return

    if args.command == "plan":
        plan = workflow.build_execution_plan()
        workflow.print_plan()
        return

    if args.command == "run":
        plan = workflow.build_execution_plan()
        workflow.print_plan()

        if args.stages == "all":
            stages_to_run = [d for d in plan if d.should_run]
        else:
            target = args.stages
            stages_to_run = [d for d in plan if d.should_run and target in d.stage]

        if not stages_to_run:
            print("无需执行任何阶段")
            return

        print(f"即将执行 {len(stages_to_run)} 个阶段...\n")

        for decision in stages_to_run:
            print(f"\n执行 [{decision.stage}]...")
            result = workflow.execute_stage(decision.stage)

            if result.get("status") == "success":
                print(f"[{decision.stage}] 完成 ({result.get('elapsed_sec', 0):.1f}s)")
            else:
                print(f"[{decision.stage}] 失败: {result.get('error', result.get('stderr', '未知错误'))}")
                break

        workflow.save_execution_report()
        print(f"\n执行完成，报告已保存至 {args.vehicle_id}/.workflow_execution_log.json")


if __name__ == "__main__":
    main()
