"""数据读取：Excel优先，SQLite回退."""

import os
import sqlite3
from pathlib import Path

import openpyxl


def _read_ripple_excel(excel_path: str) -> list[dict]:
    """读取纹波Excel，返回记录列表."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Detailed Results"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        records.append({
            "component": row[1],
            "condition_id": row[3],
            "condition_name": row[4],
            "soc_level": row[5],
            "time_vpp": row[7],
            "image_path": row[12],
        })
    return records


def _read_ripple_sqlite(db_path: str) -> list[dict]:
    """读取纹波SQLite，返回记录列表."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("""
        SELECT
            tr.component_code AS component,
            tr.condition_id,
            c.condition_name,
            c.soc_level,
            tr.time_vpp,
            tr.image_path
        FROM test_results tr
        JOIN conditions c ON tr.condition_id = c.condition_id
    """)
    records = [dict(row) for row in c.fetchall()]
    conn.close()
    return records


def _read_slope_excel(excel_path: str) -> list[dict]:
    """读取斜率Excel，返回记录列表."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Detailed Results"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        records.append({
            "component": row[1],
            "condition_id": row[3],
            "condition_name": row[4],
            "soc_level": row[5],
            "slope_max_abs": row[8],
            "image_path": row[9],
        })
    return records


def _read_slope_sqlite(db_path: str, vehicle_id: str, base_dir: str) -> list[dict]:
    """读取斜率SQLite，返回记录列表.

    slope_results表没有image_path，需根据文件规律推断.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("""
        SELECT
            sr.component_code AS component,
            sr.condition_id,
            c.condition_name,
            c.soc_level,
            sr.slope_max_abs
        FROM slope_results sr
        JOIN conditions c ON sr.condition_id = c.condition_id
    """)
    rows = c.fetchall()
    conn.close()

    records = []
    for row in rows:
        rec = dict(row)
        # 推断图片路径: {base_dir}/{vehicle_id}/{vehicle_id}_SLOPE/{component}/{condition_id}_{component}.png
        cid = rec.get("condition_id", "")
        comp = rec.get("component", "")
        inferred_path = os.path.join(
            base_dir, vehicle_id,
            f"{vehicle_id}_SLOPE", comp,
            f"{cid}_{comp}.png"
        )
        if os.path.exists(inferred_path):
            rec["image_path"] = inferred_path
        else:
            rec["image_path"] = None
        records.append(rec)
    return records


def load_ripple_data(vehicle_id: str, base_dir: str) -> list[dict]:
    """加载纹波数据，Excel优先，SQLite回退."""
    ripple_dir = os.path.join(
        base_dir, vehicle_id,
        f"{vehicle_id}_RIPPLE",
        f"{vehicle_id}_RIPPLE_output"
    )
    excel_path = os.path.join(ripple_dir, f"{vehicle_id}_RIPPLE_summary.xlsx")
    db_path = os.path.join(ripple_dir, f"{vehicle_id}_RIPPLE.db")

    # Excel优先
    if os.path.exists(excel_path):
        try:
            records = _read_ripple_excel(excel_path)
            if records:
                return records
        except Exception:
            pass

    # SQLite回退
    if os.path.exists(db_path):
        try:
            records = _read_ripple_sqlite(db_path)
            if records:
                return records
        except Exception:
            pass

    raise FileNotFoundError(f"找不到{vehicle_id}纹波数据: {excel_path} 或 {db_path}")


def load_slope_data(vehicle_id: str, base_dir: str) -> list[dict]:
    """加载斜率数据，Excel优先，SQLite回退."""
    slope_dir = os.path.join(
        base_dir, vehicle_id,
        f"{vehicle_id}_SLOPE",
        f"{vehicle_id}_SLOPE_output"
    )
    excel_path = os.path.join(slope_dir, f"{vehicle_id}_SLOPE_summary.xlsx")
    db_path = os.path.join(slope_dir, f"{vehicle_id}_SLOPE.db")

    # Excel优先
    if os.path.exists(excel_path):
        try:
            records = _read_slope_excel(excel_path)
            if records:
                return records
        except Exception:
            pass

    # SQLite回退
    if os.path.exists(db_path):
        try:
            records = _read_slope_sqlite(db_path, vehicle_id, base_dir)
            if records:
                return records
        except Exception:
            pass

    raise FileNotFoundError(f"找不到{vehicle_id}斜率数据: {excel_path} 或 {db_path}")


def filter_by_component(records: list[dict], component_code: str) -> list[dict]:
    """按组件通道过滤记录."""
    return [r for r in records if r.get("component") == component_code]


def filter_by_soc(records: list[dict], soc_level: str) -> list[dict]:
    """按SOC区间过滤记录."""
    return [r for r in records if r.get("soc_level") == soc_level]


def get_components_from_db(db_path: str) -> list[str]:
    """从数据库获取所有组件通道代码."""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("SELECT component_code FROM components")
    components = [row[0] for row in c.fetchall()]
    conn.close()
    return components


def get_components_from_excel(excel_path: str) -> list[str]:
    """从Excel获取所有组件通道代码."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Detailed Results"]
    comps = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            comps.add(row[1])
    return sorted(comps)


def detect_ripple_components(vehicle_id: str, base_dir: str) -> list[str]:
    """自动检测纹波组件通道列表."""
    ripple_dir = os.path.join(
        base_dir, vehicle_id,
        f"{vehicle_id}_RIPPLE",
        f"{vehicle_id}_RIPPLE_output"
    )
    excel_path = os.path.join(ripple_dir, f"{vehicle_id}_RIPPLE_summary.xlsx")
    db_path = os.path.join(ripple_dir, f"{vehicle_id}_RIPPLE.db")

    if os.path.exists(excel_path):
        try:
            return get_components_from_excel(excel_path)
        except Exception:
            pass
    if os.path.exists(db_path):
        try:
            return get_components_from_db(db_path)
        except Exception:
            pass
    return []


def detect_slope_components(vehicle_id: str, base_dir: str) -> list[str]:
    """自动检测斜率组件通道列表."""
    slope_dir = os.path.join(
        base_dir, vehicle_id,
        f"{vehicle_id}_SLOPE",
        f"{vehicle_id}_SLOPE_output"
    )
    excel_path = os.path.join(slope_dir, f"{vehicle_id}_SLOPE_summary.xlsx")
    db_path = os.path.join(slope_dir, f"{vehicle_id}_SLOPE.db")

    if os.path.exists(excel_path):
        try:
            return get_components_from_excel(excel_path)
        except Exception:
            pass
    if os.path.exists(db_path):
        try:
            return get_components_from_db(db_path)
        except Exception:
            pass
    return []
