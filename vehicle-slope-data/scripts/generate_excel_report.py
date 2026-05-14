#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
车辆斜率数据Excel报告生成器
根据处理后的JSON数据生成Excel报告

使用方法:
    python generate_excel_report.py --input-json data.json --output-excel report.xlsx
    python generate_excel_report.py --vehicle-folder V0001_SLOPE --output-excel report.xlsx

功能:
    1. 生成车辆信息工作表
    2. 生成组件摘要工作表
    3. 生成详细结果工作表（9列斜率数据）
"""

import argparse
import json
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
import sys

# Add parent directory to path to import config
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import SlopeConfigManager, get_slope_config_manager


class ConfigurableExcelGenerator:
    """Configuration-driven Excel report generator"""
    
    def __init__(self, config_manager: Optional[SlopeConfigManager] = None):
        self.config = config_manager or get_slope_config_manager()
        self.template = self.config.load('slope/excel_template')
        self.styles = self.config.load('common/styles')
    
    def generate(self, data: Dict, output_path: str):
        """Generate Excel report using configuration"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Generate sheets from template
        for sheet_config in self.template.get('sheets', []):
            self._create_sheet(wb, sheet_config, data)
        
        # Save workbook
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return output_path
    
    def _create_sheet(self, wb, sheet_config: Dict, data: Dict):
        """Create a sheet based on template configuration"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        sheet_name = sheet_config.get('name', 'Sheet')
        ws = wb.create_sheet(title=sheet_name)
        
        sheet_type = sheet_config.get('type')
        
        if sheet_type == 'vehicle_info':
            self._fill_vehicle_info_sheet(ws, sheet_config, data)
        elif sheet_type == 'component_summary':
            self._fill_component_summary_sheet(ws, sheet_config, data)
        elif sheet_type == 'detailed_results':
            self._fill_detailed_results_sheet(ws, sheet_config, data)
        
        # Apply column widths
        for col_idx, width in enumerate(sheet_config.get('column_widths', []), start=1):
            col_letter = self._get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
    
    def _get_column_letter(self, idx: int) -> str:
        """Convert column index to letter (1=A, 2=B, ...)"""
        result = ""
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            result = chr(65 + remainder) + result
        return result
    
    def _fill_vehicle_info_sheet(self, ws, sheet_config: Dict, data: Dict):
        """Fill vehicle information sheet - Unified format with ripple-data skill"""
        from openpyxl.styles import Font, Alignment, PatternFill

        vehicle = data.get('vehicle', {})
        vehicle_info = vehicle.get('vehicle_info', {})

        # Add header row - "Parameter" and "Value" (matching ripple-data format)
        header_fill = PatternFill(start_color='D5E8F0', end_color='D5E8F0', fill_type='solid')

        cell = ws.cell(row=1, column=1, value='Parameter')
        cell.font = Font(bold=True, color='000000')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = ws.cell(row=1, column=2, value='Value')
        cell.font = Font(bold=True, color='000000')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Standard fields (same order as ripple-data)
        standard_fields = [
            ('Vehicle ID', vehicle.get('vehicle_id', 'Unknown')),
            ('Vehicle Model', vehicle_info.get('vehicle_model', 'Unknown')),
            ('Manufacturer', vehicle_info.get('manufacturer', 'Unknown')),
            ('Length (mm)', vehicle_info.get('length_mm', '')),
            ('Width (mm)', vehicle_info.get('width_mm', '')),
            ('Height (mm)', vehicle_info.get('height_mm', '')),
        ]

        # Add extra fields from vehicle_info (same exclusions as ripple-data)
        excluded_keys = {
            'vehicle_id', 'vehicle_model', 'manufacturer',
            'length_mm', 'width_mm', 'height_mm',
        }

        extra_fields = []
        for key, value in vehicle_info.items():
            if key not in excluded_keys and value and str(value).strip():
                # Convert display name: title case for underscored keys
                if '_' in key and not any(c in key for c in '()（）'):
                    display_key = key.replace('_', ' ').title()
                else:
                    display_key = key
                extra_fields.append((display_key, value))

        # Combine all fields
        all_fields = standard_fields + extra_fields

        # Fill data starting from row 2 (after header)
        for row_idx, (key, value) in enumerate(all_fields, start=2):
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=value)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center')
    
    def _fill_component_summary_sheet(self, ws, sheet_config: Dict, data: Dict):
        """Fill component summary sheet"""
        headers = [col['header'] for col in sheet_config.get('columns', [])]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Generate data
        summary_data = []
        for comp_code, comp_data in data.get('components', {}).items():
            slopes = [cond['slope'] for cond in comp_data['conditions'].values()]
            max_vals = [s['max_value'] for s in slopes if s['max_value'] is not None]
            min_vals = [s['min_value'] for s in slopes if s['min_value'] is not None]
            max_abs_vals = [s['max_abs_value'] for s in slopes if s['max_abs_value'] is not None]
            
            summary_data.append({
                'Component Code': comp_code,
                'Component Name': comp_data['component_name'],
                'Unit': comp_data['unit'],
                'Conditions Count': comp_data['conditions_count'],
                'Max Slope (V/s)': max(max_vals) if max_vals else None,
                'Min Slope (V/s)': min(min_vals) if min_vals else None,
                'Max Abs Slope (V/s)': max(max_abs_vals) if max_abs_vals else None
            })
        
        # Write data rows
        for row_idx, row_data in enumerate(summary_data, start=2):
            for col_idx, col_config in enumerate(sheet_config.get('columns', []), start=1):
                header = col_config['header']
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _fill_detailed_results_sheet(self, ws, sheet_config: Dict, data: Dict):
        """Fill detailed results sheet"""
        headers = [col['header'] for col in sheet_config.get('columns', [])]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Generate data
        results_data = []
        seq_num = 1
        for comp_code, comp_data in data.get('components', {}).items():
            unit = comp_data['unit']
            for cond_id, cond_data in comp_data['conditions'].items():
                results_data.append({
                    'No.': seq_num,
                    'Component': comp_code,
                    'Unit': unit,
                    'Condition ID': cond_id,
                    'Condition Name': cond_data['condition_name'],
                    'SOC Level': cond_data['soc_level'],
                    'Slope Max (V/s)': cond_data['slope']['max_value'],
                    'Slope Min (V/s)': cond_data['slope']['min_value'],
                    'Slope Max Abs (V/s)': cond_data['slope']['max_abs_value'],
                    'Image Path': cond_data.get('image_path', '')
                })
                seq_num += 1
        
        # Write data rows
        for row_idx, row_data in enumerate(results_data, start=2):
            for col_idx, col_config in enumerate(sheet_config.get('columns', []), start=1):
                header = col_config['header']
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply alignment if specified
                align = col_config.get('align')
                if align:
                    cell.alignment = Alignment(horizontal=align, vertical='center')


def generate_excel_report(data: Dict, output_path: str, use_config: bool = True):
    """
    Generate Excel report
    
    Parameters:
        data: Processed data dictionary
        output_path: Output Excel file path
        use_config: Whether to use configuration-driven generation (default: True)
    """
    if use_config:
        generator = ConfigurableExcelGenerator()
        return generator.generate(data, output_path)
    
    # Fallback to legacy implementation
    output_file = Path(output_path)
    
    # 确保输出目录存在
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 工作表1: 车辆信息 (竖列格式)
        vehicle_info = data['vehicle']['vehicle_info']
        vehicle_data = [{'Parameter': key, 'Value': value} for key, value in vehicle_info.items()]
        vehicle_df = pd.DataFrame(vehicle_data)
        vehicle_df.to_excel(writer, sheet_name='Vehicle Information', index=False)
        
        # 工作表2: 组件摘要
        summary_data = []
        for comp_code, comp_data in data['components'].items():
            # 计算斜率统计
            slopes = [cond['slope'] for cond in comp_data['conditions'].values()]
            max_vals = [s['max_value'] for s in slopes if s['max_value'] is not None]
            min_vals = [s['min_value'] for s in slopes if s['min_value'] is not None]
            max_abs_vals = [s['max_abs_value'] for s in slopes if s['max_abs_value'] is not None]
            
            summary_data.append({
                'Component Code': comp_code,
                'Component Name': comp_data['component_name'],
                'Unit': comp_data['unit'],
                'Conditions Count': comp_data['conditions_count'],
                'Max Slope (V/s)': max(max_vals) if max_vals else None,
                'Min Slope (V/s)': min(min_vals) if min_vals else None,
                'Max Abs Slope (V/s)': max(max_abs_vals) if max_abs_vals else None
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Component Summary', index=False)
        
        # 工作表3: 详细结果
        results_data = []
        seq_num = 1
        for comp_code, comp_data in data['components'].items():
            unit = comp_data['unit']
            for cond_id, cond_data in comp_data['conditions'].items():
                results_data.append({
                    'No.': seq_num,
                    'Component': comp_code,
                    'Unit': unit,
                    'Condition ID': cond_id,
                    'Condition Name': cond_data['condition_name'],
                    'SOC Level': cond_data['soc_level'],
                    'Slope Max (V/s)': cond_data['slope']['max_value'],
                    'Slope Min (V/s)': cond_data['slope']['min_value'],
                    'Slope Max Abs (V/s)': cond_data['slope']['max_abs_value'],
                    'Image Path': cond_data.get('image_path', '')
                })
                seq_num += 1
        
        results_df = pd.DataFrame(results_data)
        results_df.to_excel(writer, sheet_name='Detailed Results', index=False)
        
        # 调整列宽
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            if sheet_name == 'Vehicle Information':
                worksheet.column_dimensions['A'].width = 25
                worksheet.column_dimensions['B'].width = 30
            elif sheet_name == 'Component Summary':
                worksheet.column_dimensions['A'].width = 18
                worksheet.column_dimensions['B'].width = 35
                worksheet.column_dimensions['C'].width = 10
                worksheet.column_dimensions['D'].width = 18
                for col in ['E', 'F', 'G']:
                    worksheet.column_dimensions[col].width = 20
            elif sheet_name == 'Detailed Results':
                worksheet.column_dimensions['A'].width = 8
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 8
                worksheet.column_dimensions['D'].width = 30
                worksheet.column_dimensions['E'].width = 25
                worksheet.column_dimensions['F'].width = 12
                for col in ['G', 'H', 'I']:
                    worksheet.column_dimensions[col].width = 20


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='车辆斜率数据Excel报告生成器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  # 从JSON文件生成
  python generate_excel_report.py --input-json V0001_SLOPE_data.json --output-excel V0001_SLOPE_summary.xlsx
  
  # 自动从车辆文件夹读取JSON
  python generate_excel_report.py --vehicle-folder V0001_SLOPE --output-excel V0001_SLOPE_summary.xlsx
        '''
    )
    
    parser.add_argument(
        '--input-json', '-i',
        help='输入JSON文件路径'
    )
    
    parser.add_argument(
        '--vehicle-folder', '-f',
        help='车辆文件夹路径（自动读取SKILL_output/{VehID}_data.json）'
    )
    
    parser.add_argument(
        '--output-excel', '-o',
        required=True,
        help='输出Excel文件路径'
    )
    
    args = parser.parse_args()
    
    # 确定输入文件
    if args.input_json:
        input_path = Path(args.input_json)
    elif args.vehicle_folder:
        vehicle_folder = Path(args.vehicle_folder)
        folder_name = vehicle_folder.name
        
        # 提取vehicle_id
        vehicle_id = folder_name[:-6] if folder_name.endswith('_SLOPE') else folder_name
        
        input_path = vehicle_folder / f"{vehicle_id}_SLOPE_output" / f"{vehicle_id}_SLOPE_data.json"
    else:
        print("错误: 必须指定 --input-json 或 --vehicle-folder")
        return 1
    
    # 检查输入文件
    if not input_path.exists():
        print(f"错误: 输入文件不存在: {input_path}")
        return 1
    
    # 读取JSON数据
    try:
        with open(input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"错误: 无法读取JSON文件: {e}")
        return 1
    
    # 生成Excel
    try:
        generate_excel_report(data, args.output_excel)
        print(f"[OK] Excel report generated: {args.output_excel}")
        print(f"     Sheets: Vehicle Information, Component Summary, Detailed Results")
        return 0
    except Exception as e:
        print(f"Error: Failed to generate Excel: {e}")
        return 1


if __name__ == '__main__':
    exit(main())
